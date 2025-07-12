from pathlib import Path

import openpyxl
import pandas as pd

from pa_core.reporting import export_to_excel


def test_export_to_excel_sheets(tmp_path: Path):
    inputs = {"a": 1, "b": 2}
    summary = pd.DataFrame({"Base": [0.1]})
    raw = {"Base": pd.DataFrame([[0.1, 0.2]], columns=[0, 1])}
    file_path = tmp_path / "out.xlsx"
    export_to_excel(inputs, summary, raw, filename=str(file_path))
    wb = openpyxl.load_workbook(file_path)
    assert set(wb.sheetnames) == {"Inputs", "Summary", "Base"}


def test_export_to_excel_pivot(tmp_path: Path):
    inputs = {"x": 1}
    summary = pd.DataFrame({"Total": [0.2]})
    raw = {
        "Base": pd.DataFrame([[0.1, 0.2]], columns=[0, 1]),
        "Ext": pd.DataFrame([[0.3, 0.4]], columns=[0, 1]),
    }
    file_path = tmp_path / "pivot.xlsx"
    export_to_excel(inputs, summary, raw, filename=str(file_path), pivot=True)
    wb = openpyxl.load_workbook(file_path)
    assert set(wb.sheetnames) == {"Inputs", "Summary", "AllReturns"}
    ws = wb["AllReturns"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["Sim", "Month", "Agent", "Return"]
