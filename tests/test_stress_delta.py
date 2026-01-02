import io

import openpyxl
import pandas as pd
import pytest

from pa_core.reporting.stress_delta import (
    build_delta_table,
    build_stress_workbook,
    format_delta_table_text,
)


def test_build_delta_table_includes_total_row():
    base = pd.DataFrame(
        {
            "Agent": ["Base", "ExternalPA", "Total"],
            "terminal_AnnReturn": [0.05, 0.07, 0.12],
            "monthly_BreachCountPath0": [1, 2, 3],
        }
    )
    stressed = pd.DataFrame(
        {
            "Agent": ["Base", "ExternalPA", "Total"],
            "terminal_AnnReturn": [0.02, 0.06, 0.08],
            "monthly_BreachCountPath0": [3, 5, 5],
        }
    )

    delta_df = build_delta_table(base, stressed)

    assert "Total" in delta_df["Agent"].tolist()
    assert "Base" in delta_df["Agent"].tolist()
    assert delta_df["Agent"].iloc[-1] == "Total"

    total_row = delta_df[delta_df["Agent"] == "Total"].iloc[0]
    assert total_row["terminal_AnnReturn"] == pytest.approx(-0.04)
    assert total_row["monthly_BreachCountPath0"] == pytest.approx(2)


def test_build_stress_workbook_contains_expected_sheets():
    base = pd.DataFrame({"Agent": ["Base"], "terminal_AnnReturn": [0.05], "monthly_BreachCountPath0": [1]})
    stressed = pd.DataFrame({"Agent": ["Base"], "terminal_AnnReturn": [0.02], "monthly_BreachCountPath0": [3]})
    delta_df = build_delta_table(base, stressed)
    config_diff = pd.DataFrame({"Parameter": ["mu_H"], "Base": [0.01], "Stressed": [0]})

    data = build_stress_workbook(base, stressed, delta_df, config_diff)
    wb = openpyxl.load_workbook(io.BytesIO(data))

    assert "BaseSummary" in wb.sheetnames
    assert "StressedSummary" in wb.sheetnames
    assert "Delta" in wb.sheetnames
    assert "ConfigDiff" in wb.sheetnames


def test_format_delta_table_text_adds_signs():
    delta_df = pd.DataFrame(
        {"Agent": ["Total"], "terminal_AnnReturn": [-0.03], "monthly_AnnVol": [0.01], "monthly_BreachCountPath0": [2]}
    )
    formatted = format_delta_table_text(delta_df)

    assert formatted.loc[0, "terminal_AnnReturn"] == "-3.00%"
    assert formatted.loc[0, "monthly_AnnVol"] == "+1.00%"
    assert formatted.loc[0, "monthly_BreachCountPath0"] == "+2.0000"
