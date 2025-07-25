#!/usr/bin/env python
# coding: utf-8

# **_🚨 IMPORTANT: BEFORE RUNNING THE PROGRAM, SAVE THIS SHEET AS A CSV FILE! 🚨_**
# 
# import pandas as pd
# 
# rows = [
#     ("Parameter", "Value", "Notes"),  # Header row
#     ("Analysis mode", "", "Choose one: capital, returns, alpha_shares, vol_mult"),
#     # Capital mode
#     ("Max external combined (%)", "", "capital mode: max % for (External PA + Active Ext)"),
#     ("External step size (%)", "", "capital mode: increment for external %"),
#     # Fixed capital (for returns / alpha_shares / vol_mult)
#     ("External PA capital (mm)", "", "returns/alpha_shares/vol_mult modes: allocation in mm"),
#     ("Active Extension capital (mm)", "", ""),
#     ("Internal PA capital (mm)", "", ""),
#     # Returns mode ranges (min/max/step)
#     ("In-House return min (%)", "", "returns mode"),
#     ("In-House return max (%)", "", ""),
#     ("In-House return step (%)", "", ""),
#     ("In-House vol min (%)", "", "returns mode"),
#     ("In-House vol max (%)", "", ""),
#     ("In-House vol step (%)", "", ""),
#     ("Alpha-Extension return min (%)", "", "returns mode"),
#     ("Alpha-Extension return max (%)", "", ""),
#     ("Alpha-Extension return step (%)", "", ""),
#     ("Alpha-Extension vol min (%)", "", "returns mode"),
#     ("Alpha-Extension vol max (%)", "", ""),
#     ("Alpha-Extension vol step (%)", "", ""),
#     ("External return min (%)", "", "returns mode"),
#     ("External return max (%)", "", ""),
#     ("External return step (%)", "", ""),
#     ("External vol min (%)", "", "returns mode"),
#     ("External vol max (%)", "", ""),
#     ("External vol step (%)", "", ""),
#     # Alpha_Shares mode ranges
#     ("External PA α fraction min (%)", "", "alpha_shares mode"),
#     ("External PA α fraction max (%)", "", ""),
#     ("External PA α fraction step (%)", "", ""),
#     ("Active share min (%)", "", "alpha_shares mode"),
#     ("Active share max (%)", "", ""),
#     ("Active share step (%)", "", ""),
#     # Vol_Mult mode ranges
#     ("SD multiple min", "", "vol_mult mode"),
#     ("SD multiple max", "", ""),
#     ("SD multiple step", "", ""),
#     # Financing & overrides
#     ("Annual financing mean (%)", "", "global financing"),
#     ("Annual financing vol (%)", "", ""),
#     ("Monthly spike probability", "", ""),
#     ("Spike size (σ × multiplier)", "", ""),
#     ("Internal financing mean (%)", "", "internal override"),
#     ("Internal financing vol (%)", "", ""),
#     ("Internal monthly spike probability", "", ""),
#     ("Internal spike size (σ × multiplier)", "", ""),
#     ("External PA financing mean (%)", "", "ext PA override"),
#     ("External PA financing vol (%)", "", ""),
#     ("External PA monthly spike probability", "", ""),
#     ("External PA spike size (σ × multiplier)", "", ""),
#     ("Active Extension financing mean (%)", "", "active ext override"),
#     ("Active Extension financing vol (%)", "", ""),
#     ("Active Extension monthly spike probability", "", ""),
#     ("Active Extension spike size (σ × multiplier)", "", ""),
#     # Alpha stream fallbacks
#     ("In-House annual return (%)", "", "fallback if no range"),
#     ("In-House annual vol (%)", "", ""),
#     ("Alpha-Extension annual return (%)", "", ""),
#     ("Alpha-Extension annual vol (%)", "", ""),
#     ("External annual return (%)", "", ""),
#     ("External annual vol (%)", "", ""),
#     # Correlations
#     ("Corr index–In-House", "", ""),
#     ("Corr index–Alpha-Extension", "", ""),
#     ("Corr index–External", "", ""),
#     ("Corr In-House–Alpha-Extension", "", ""),
#     ("Corr In-House–External", "", ""),
#     ("Corr Alpha-Extension–External", "", ""),
#     # Buffer multiple & total capital
#     ("Buffer multiple", "", "cash‐buffer multiple"),
#     ("Total fund capital (mm)", "", "total fund size in mm")
# ]
# 
# # Build DataFrame and save
# df_template = pd.DataFrame(rows, columns=["Parameter", "Value", "Notes"])
# file_path = "/mnt/data/parameters_template.xlsx"
# df_template.to_excel(file_path, index=False)
# print(f"Created template: {file_path}")
# 

# In[12]:


# portable_alpha_model.py

import sys
import csv
import numpy as np
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import openpyxl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

# =============================================================================
# 1. MAPPING: User-friendly labels → Internal variable names
# =============================================================================

LABEL_MAP = {
    "Analysis mode":                           "analysis_mode",
    # Capital mode inputs
    "Max external combined (%)":               "max_external_combined_percent",
    "External step size (%)":                  "external_step_size_percent",
    # Fixed-capital (for non-capital modes)
    "External PA capital (mm)":                "external_pa_capital",
    "Active Extension capital (mm)":           "active_ext_capital",
    "Internal PA capital (mm)":                "internal_pa_capital",
    # Returns mode ranges
    "In-House return min (%)":                 "mu_H_min",
    "In-House return max (%)":                 "mu_H_max",
    "In-House return step (%)":                "mu_H_step",
    "In-House vol min (%)":                    "sigma_H_min",
    "In-House vol max (%)":                    "sigma_H_max",
    "In-House vol step (%)":                   "sigma_H_step",
    "Alpha-Extension return min (%)":          "mu_E_min",
    "Alpha-Extension return max (%)":          "mu_E_max",
    "Alpha-Extension return step (%)":         "mu_E_step",
    "Alpha-Extension vol min (%)":             "sigma_E_min",
    "Alpha-Extension vol max (%)":             "sigma_E_max",
    "Alpha-Extension vol step (%)":            "sigma_E_step",
    "External return min (%)":                 "mu_M_min",
    "External return max (%)":                 "mu_M_max",
    "External return step (%)":                "mu_M_step",
    "External vol min (%)":                    "sigma_M_min",
    "External vol max (%)":                    "sigma_M_max",
    "External vol step (%)":                   "sigma_M_step",
    # Alpha_Shares mode ranges
    "External PA α fraction min (%)":          "external_pa_alpha_frac_min",
    "External PA α fraction max (%)":          "external_pa_alpha_frac_max",
    "External PA α fraction step (%)":         "external_pa_alpha_frac_step",
    "Active share min (%)":                    "active_share_min",
    "Active share max (%)":                    "active_share_max",
    "Active share step (%)":                   "active_share_step",
    # Vol_Mult mode range
    "SD multiple min":                         "sd_of_vol_mult_min",
    "SD multiple max":                         "sd_of_vol_mult_max",
    "SD multiple step":                        "sd_of_vol_mult_step",
    # Financing & bucket overrides
    "Annual financing mean (%)":               "financing_mean_annual",
    "Annual financing vol (%)":                "financing_vol_annual",
    "Monthly spike probability":               "spike_prob",
    "Spike size (σ × multiplier)":             "spike_factor",
    "Internal financing mean (%)":             "internal_financing_mean_annual",
    "Internal financing vol (%)":              "internal_financing_vol_annual",
    "Internal monthly spike probability":      "internal_spike_prob",
    "Internal spike size (σ × multiplier)":    "internal_spike_factor",
    "External PA financing mean (%)":          "ext_pa_financing_mean_annual",
    "External PA financing vol (%)":           "ext_pa_financing_vol_annual",
    "External PA monthly spike probability":   "ext_pa_spike_prob",
    "External PA spike size (σ × multiplier)": "ext_pa_spike_factor",
    "Active Extension financing mean (%)":     "act_ext_financing_mean_annual",
    "Active Extension financing vol (%)":      "act_ext_financing_vol_annual",
    "Active Extension monthly spike probability":"act_ext_spike_prob",
    "Active Extension spike size (σ × multiplier)":"act_ext_spike_factor",
    # Fallback alpha stream defaults
    "In-House annual return (%)":              "mu_H",
    "In-House annual vol (%)":                 "sigma_H",
    "Alpha-Extension annual return (%)":       "mu_E",
    "Alpha-Extension annual vol (%)":          "sigma_E",
    "External annual return (%)":              "mu_M",
    "External annual vol (%)":                 "sigma_M",
    # Correlations
    "Corr index–In-House":                     "rho_idx_H",
    "Corr index–Alpha-Extension":              "rho_idx_E",
    "Corr index–External":                     "rho_idx_M",
    "Corr In-House–Alpha-Extension":           "rho_H_E",
    "Corr In-House–External":                  "rho_H_M",
    "Corr Alpha-Extension–External":            "rho_E_M",
    # Other risk controls
    "Buffer multiple":                         "buffer_multiple",
    "Total fund capital (mm)":                 "total_fund_capital",
    # Visualization toggles
    "Plot heatmap":                            "plot_heatmap",
    "Plot line":                               "plot_line",
    "Plot boxplot":                            "plot_boxplot",
    "Plot scatter":                            "plot_scatter",
    "Plot time series":                        "plot_time_series",
    "Plot histogram":                          "plot_histogram",
    "Plot surface":                            "plot_surface",
}

# =============================================================================
# 2. FILE‐PICKER FOR CSV SELECTION
# =============================================================================

def select_csv_file():
    """
    Pop up a file‐picker dialog so the user can choose a CSV file.
    Returns a pathlib.Path to the selected file.
    Raises FileNotFoundError if the user cancels.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select CSV File",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    root.destroy()
    if not file_path:
        raise FileNotFoundError("No file selected.")
    return Path(file_path)

# =============================================================================
# 3. LOAD PARAMETERS USING MAPPING
# =============================================================================

def load_parameters(csv_filepath, label_map):
    """
    Read a CSV that may have leading instruction rows, then a header row "Parameter,Value".
    Skip all rows until the header, then parse friendly labels → internal names via label_map.
    Returns a dict {internal_var_name: parsed_value}.
    """
    params = {}
    lines = Path(csv_filepath).read_text(encoding="utf-8").splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith("Parameter,"):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"No header row starting with 'Parameter,' found in {csv_filepath}")

    header_and_data = lines[header_idx:]
    reader = csv.DictReader(header_and_data)

    for row in reader:
        friendly_key = row.get("Parameter", "").strip()
        if not friendly_key or friendly_key not in label_map:
            continue

        internal_key = label_map[friendly_key]
        raw_val = row.get("Value", "").strip()

        # If semicolon is present, parse as a list (legacy fallback)
        if ";" in raw_val:
            parts = [p.strip() for p in raw_val.split(";") if p.strip() != ""]
            parsed_list = []
            for p in parts:
                try:
                    if "." in p:
                        parsed_list.append(float(p))
                    else:
                        parsed_list.append(int(p))
                except ValueError:
                    parsed_list.append(p)
            params[internal_key] = parsed_list
        else:
            # Try int → float → string
            try:
                params[internal_key] = int(raw_val)
            except ValueError:
                try:
                    params[internal_key] = float(raw_val)
                except ValueError:
                    # Could be blank string; store as-is
                    params[internal_key] = raw_val

    return params

# =============================================================================
# 4. UTILITY: “SAFE GET” FOR NUMERIC PARAMETERS
# =============================================================================

def get_num(raw_params, key, default):
    """
    Return raw_params[key] if it's already int or float; otherwise default.
    """
    v = raw_params.get(key, None)
    if isinstance(v, (int, float)):
        return v
    return default

# =============================================================================
# 5. HELPER FUNCTIONS TO BUILD RANGES OR FALLBACK TO MIDPOINT
# =============================================================================

def build_range(key_base, default_midpoint):
    """
    If raw_params contains key_base_min & key_base_max (and optionally key_base_step),
    return a list of decimals from min→max (in steps).
    Otherwise, if key_base_list exists (legacy semicolon list), return that.
    Else, return [default_midpoint].
    We divide by 100 because these are “percent” inputs.
    """
    k_min  = get_num(raw_params, f"{key_base}_min", None)
    k_max  = get_num(raw_params, f"{key_base}_max", None)
    k_step = get_num(raw_params, f"{key_base}_step", None)

    if (k_min is not None) and (k_max is not None):
        step = k_step if (k_step is not None) else (k_max - k_min)
        if step <= 0:
            raise RuntimeError(f"Step for '{key_base}' must be positive.")
        start = k_min / 100.0
        stop  = k_max / 100.0
        stepd = step / 100.0
        arr = np.arange(start, stop + 1e-9, stepd)
        return list(arr)

    # Legacy semicolon‐list fallback
    flat_list = raw_params.get(f"{key_base}_list", None)
    if isinstance(flat_list, list):
        return flat_list

    return [default_midpoint]

def build_range_int(key_base, default_midpoint):
    """
    Like build_range but for integer inputs (e.g. SD multiple).
    If key_base_min & key_base_max exist, return list(range(min, max+1, step)).
    Else if key_base_list exists, return it. Else [default_midpoint].
    """
    k_min  = get_num(raw_params, f"{key_base}_min", None)
    k_max  = get_num(raw_params, f"{key_base}_max", None)
    k_step = get_num(raw_params, f"{key_base}_step", None)

    if (k_min is not None) and (k_max is not None):
        step = k_step if (k_step is not None) else (k_max - k_min)
        if step <= 0:
            raise RuntimeError(f"Step for '{key_base}' must be positive.")
        return list(range(k_min, k_max + 1, step))

    flat_list = raw_params.get(f"{key_base}_list", None)
    if isinstance(flat_list, list):
        return flat_list

    return [default_midpoint]

# =============================================================================
# 6. HELPER TO LOAD INDEX RETURNS
# =============================================================================

def load_index_returns(csv_path):
    """
    Load a CSV of monthly index returns into a pandas Series.
    Expects columns: "Date" and either "Monthly_TR" or "Return".
    Returns a pd.Series indexed by Date (datetime).
    """
    csv_path = Path(csv_path)
    if not csv_path.exists() or not csv_path.is_file():
        raise FileNotFoundError(f"Index CSV not found at {csv_path}")
    df = pd.read_csv(csv_path, parse_dates=["Date"])
    if "Date" not in df.columns:
        raise ValueError(f"'Date' column is missing from {csv_path}")
    if "Monthly_TR" in df.columns:
        col = "Monthly_TR"
    elif "Return" in df.columns:
        col = "Return"
    else:
        raise ValueError(f"CSV must contain 'Monthly_TR' or 'Return'; found: {df.columns.tolist()}")

    df = df.sort_values("Date").reset_index(drop=True)
    df.set_index("Date", inplace=True)
    series = df[col].dropna().copy()
    series.index = pd.to_datetime(series.index)
    return series

# =============================================================================
# 7. SIMULATION + UTILITY FUNCTIONS
# =============================================================================

def simulate_financing(T, financing_mean, financing_sigma, spike_prob, spike_factor):
    """
    Simulate a series of financing spreads f_t for T months,
    using a Normal + occasional jump model.
    """
    f = np.zeros(T)
    for t in range(T):
        base = financing_mean + np.random.normal(0, financing_sigma)
        jump = 0.0
        if np.random.rand() < spike_prob:
            jump = spike_factor * financing_sigma
        f[t] = max(base + jump, 0.0)
    return f

def build_cov_matrix(rho_idx_H, rho_idx_E, rho_idx_M,
                     rho_H_E, rho_H_M, rho_E_M,
                     idx_sigma, sigma_H, sigma_E, sigma_M):
    """
    Build the 4×4 covariance matrix for (Index, H, E, M).
    """
    sds = np.array([idx_sigma, sigma_H, sigma_E, sigma_M])
    rho = np.array([
        [1.0,       rho_idx_H, rho_idx_E, rho_idx_M],
        [rho_idx_H, 1.0,       rho_H_E,   rho_H_M],
        [rho_idx_E, rho_H_E,   1.0,       rho_E_M],
        [rho_idx_M, rho_H_M,   rho_E_M,   1.0    ]
    ])
    return np.outer(sds, sds) * rho

def simulate_alpha_streams(T, cov, mu_idx, mu_H, mu_E, mu_M):
    """
    Simulate T joint observations of (Index_return, H, E, M)
    from a multivariate Normal with given means and covariance.
    Returns shape (T, 4).
    """
    means = np.array([mu_idx, mu_H, mu_E, mu_M])
    return np.random.multivariate_normal(means, cov, size=T)

def export_to_excel(inputs_dict, summary_df, raw_returns_dict, filename="Outputs.xlsx"):
    """
    Write inputs, summary, and raw returns into an Excel workbook.
    """
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 1) Inputs sheet
        df_inputs = pd.DataFrame.from_dict(inputs_dict, orient="index", columns=["Value"])
        df_inputs.index.name = "Parameter"
        df_inputs.reset_index(inplace=True)
        df_inputs.to_excel(writer, sheet_name="Inputs", index=False)

        # 2) Summary sheet
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # 3) Raw Returns sheets
        for sheet_name, df in raw_returns_dict.items():
            safe_name = sheet_name if len(sheet_name) <= 31 else sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=True)

    print(f"Exported results to {filename}")

# =============================================================================
# 8. MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    # 8.1) Prompt user to select the parameters CSV
    try:
        params_csv_path = select_csv_file()
        print(f"Parameters CSV selected: {params_csv_path}")
    except FileNotFoundError:
        raise RuntimeError("No parameter CSV selected; exiting.")

    # 8.2) Load raw parameters
    raw_params = load_parameters(params_csv_path, LABEL_MAP)

    # 8.3) “Safe get” for numeric scalars and defaults
    total_fund_capital = get_num(raw_params, "total_fund_capital", 1000)

    financing_mean_annual = get_num(raw_params, "financing_mean_annual", 0.005)
    financing_vol_annual  = get_num(raw_params, "financing_vol_annual", 0.001)
    spike_prob            = get_num(raw_params, "spike_prob", 0.02)
    spike_factor          = get_num(raw_params, "spike_factor", 2.25)

    internal_financing_mean_annual = get_num(raw_params, "internal_financing_mean_annual", financing_mean_annual)
    internal_financing_vol_annual  = get_num(raw_params, "internal_financing_vol_annual",  financing_vol_annual)
    internal_spike_prob            = get_num(raw_params, "internal_spike_prob",            spike_prob)
    internal_spike_factor          = get_num(raw_params, "internal_spike_factor",          spike_factor)

    ext_pa_financing_mean_annual = get_num(raw_params, "ext_pa_financing_mean_annual", financing_mean_annual)
    ext_pa_financing_vol_annual  = get_num(raw_params, "ext_pa_financing_vol_annual",  financing_vol_annual)
    ext_pa_spike_prob            = get_num(raw_params, "ext_pa_spike_prob",            spike_prob)
    ext_pa_spike_factor          = get_num(raw_params, "ext_pa_spike_factor",          spike_factor)

    act_ext_financing_mean_annual = get_num(raw_params, "act_ext_financing_mean_annual", financing_mean_annual)
    act_ext_financing_vol_annual  = get_num(raw_params, "act_ext_financing_vol_annual",  financing_vol_annual)
    act_ext_spike_prob            = get_num(raw_params, "act_ext_spike_prob",            spike_prob)
    act_ext_spike_factor          = get_num(raw_params, "act_ext_spike_factor",          spike_factor)

    mu_H    = get_num(raw_params, "mu_H",    0.04)
    sigma_H = get_num(raw_params, "sigma_H", 0.01)
    mu_E    = get_num(raw_params, "mu_E",    0.05)
    sigma_E = get_num(raw_params, "sigma_E", 0.02)
    mu_M    = get_num(raw_params, "mu_M",    0.03)
    sigma_M = get_num(raw_params, "sigma_M", 0.02)

    rho_idx_H = get_num(raw_params, "rho_idx_H", 0.05)
    rho_idx_E = get_num(raw_params, "rho_idx_E", 0.00)
    rho_idx_M = get_num(raw_params, "rho_idx_M", 0.00)
    rho_H_E   = get_num(raw_params, "rho_H_E",   0.10)
    rho_H_M   = get_num(raw_params, "rho_H_M",   0.10)
    rho_E_M   = get_num(raw_params, "rho_E_M",   0.00)

    buffer_multiple = get_num(raw_params, "buffer_multiple", 3.0)

    default_mu_H    = mu_H / 12
    default_sigma_H = sigma_H / 12
    default_mu_E    = mu_E / 12
    default_sigma_E = sigma_E / 12
    default_mu_M    = mu_M / 12
    default_sigma_M = sigma_M / 12

    default_ext_alpha_frac = get_num(raw_params, "external_pa_alpha_frac", 50) / 100.0
    default_act_share      = get_num(raw_params, "active_share",            50) / 100.0

    default_sd_mult = get_num(raw_params, "sd_of_vol_mult", 3)

    financing_mean  = financing_mean_annual / 12
    financing_sigma = financing_vol_annual / 12

    # 8.4) Extract visualization flags (each “Plot X” row)
    plot_heatmap_flag     = bool(str(raw_params.get("plot_heatmap", "")).strip())
    plot_line_flag        = bool(str(raw_params.get("plot_line", "")).strip())
    plot_boxplot_flag     = bool(str(raw_params.get("plot_boxplot", "")).strip())
    plot_scatter_flag     = bool(str(raw_params.get("plot_scatter", "")).strip())
    plot_time_series_flag = bool(str(raw_params.get("plot_time_series", "")).strip())
    plot_histogram_flag   = bool(str(raw_params.get("plot_histogram", "")).strip())
    plot_surface_flag     = bool(str(raw_params.get("plot_surface", "")).strip())

    # 8.5) Determine analysis_mode
    analysis_mode = str(raw_params.get("analysis_mode", "")).strip().lower()
    valid_modes = {"capital", "returns", "alpha_shares", "vol_mult"}
    if analysis_mode not in valid_modes:
        raise RuntimeError(f"Analysis mode must be one of {valid_modes}, but got '{analysis_mode}'")

    # 8.6) Prompt user to select the INDEX CSV
    print("Please select the INDEX CSV (monthly total returns).")
    try:
        INDEX_CSV_PATH = select_csv_file()
        print(f"Index CSV selected: {INDEX_CSV_PATH}")
    except FileNotFoundError:
        raise RuntimeError("Index CSV was not selected; exiting.")

    # 8.7) Load idx_series and compute reference stats
    try:
        idx_series = load_index_returns(INDEX_CSV_PATH)
        print(f"Loaded {len(idx_series)} months from the raw index CSV.")
    except Exception as e:
        raise RuntimeError(f"Failed to load index returns: {e}")

    mu_idx   = idx_series.mean()
    idx_sigma = idx_series.std(ddof=1)
    print(f"Using idx_series from {idx_series.index.min().date()} to {idx_series.index.max().date()} (n={len(idx_series)})")
    print(f"Analysis-window: μ_idx = {mu_idx:.4f}, σ_idx = {idx_sigma:.4f}")

    # 8.8) Convert bucket financing overrides to monthly decimals
    internal_financing_mean  = internal_financing_mean_annual / 12
    internal_financing_sigma = internal_financing_vol_annual / 12

    ext_pa_financing_mean  = ext_pa_financing_mean_annual / 12
    ext_pa_financing_sigma = ext_pa_financing_vol_annual / 12

    act_ext_financing_mean  = act_ext_financing_mean_annual / 12
    act_ext_financing_sigma = act_ext_financing_vol_annual / 12

    # -------------------------------------------------------------------
    # 8.9) Branch on analysis_mode
    # -------------------------------------------------------------------

    if analysis_mode == "capital":
        # --------------------------------------
        # 8.9.a) Capital mode: Sweep ext_pct→act_pct
        # --------------------------------------
        max_ext_pct  = get_num(raw_params, "max_external_combined_percent", 50) / 100.0
        ext_step_pct = get_num(raw_params, "external_step_size_percent",   1) / 100.0

        # Other “sweepable” params collapse to midpoint
        mu_H_list    = [default_mu_H]
        sigma_H_list = [default_sigma_H]
        mu_E_list    = [default_mu_E]
        sigma_E_list = [default_sigma_E]
        mu_M_list    = [default_mu_M]
        sigma_M_list = [default_sigma_M]

        ext_alpha_list = [default_ext_alpha_frac]
        act_share_list = [default_act_share]
        sd_list        = [default_sd_mult]

        all_summaries = []
        all_raw_returns = {}

        ext_range = np.arange(0.0, max_ext_pct + 1e-9, ext_step_pct)
        for ext_pct in ext_range:
            act_range = np.arange(0.0, ext_pct + 1e-9, ext_step_pct)
            for act_pct in act_range:
                E_pct = ext_pct - act_pct
                A_pct = act_pct
                Z_frac = 1.0 - ext_pct
                E_mm = E_pct * total_fund_capital
                A_mm = A_pct * total_fund_capital
                Z_mm = Z_frac * total_fund_capital

                # 1-year financing paths
                f_internal = simulate_financing(
                    12,
                    internal_financing_mean, internal_financing_sigma,
                    internal_spike_prob, internal_spike_factor
                )
                f_ext_pa = simulate_financing(
                    12,
                    ext_pa_financing_mean, ext_pa_financing_sigma,
                    ext_pa_spike_prob, ext_pa_spike_factor
                )
                f_act_ext = simulate_financing(
                    12,
                    act_ext_financing_mean, act_ext_financing_sigma,
                    act_ext_spike_prob, act_ext_spike_factor
                )

                cov_mat = build_cov_matrix(
                    rho_idx_H, rho_idx_E, rho_idx_M,
                    rho_H_E, rho_H_M, rho_E_M,
                    idx_sigma, default_sigma_H, default_sigma_E, default_sigma_M
                )

                N_SIMULATIONS = get_num(raw_params, "N_SIMULATIONS", 5000)
                N_MONTHS      = get_num(raw_params, "N_MONTHS",    12)

                sims = np.random.multivariate_normal(
                    [mu_idx, default_mu_H, default_mu_E, default_mu_M],
                    cov_mat,
                    size=(N_SIMULATIONS, N_MONTHS)
                )

                f_int_matrix    = np.tile(f_internal, (N_SIMULATIONS, 1))
                f_ext_pa_matrix = np.tile(f_ext_pa,    (N_SIMULATIONS, 1))
                f_act_ext_matrix= np.tile(f_act_ext,   (N_SIMULATIONS, 1))

                results = {
                    "Base":       np.zeros(N_SIMULATIONS),
                    "ExternalPA": np.zeros(N_SIMULATIONS),
                    "ActiveExt":  np.zeros(N_SIMULATIONS)
                }
                dates_sim = pd.date_range(
                    start=idx_series.index[-1] + pd.DateOffset(months=1),
                    periods=N_MONTHS, freq="ME"
                )
                raw_returns = {
                    "Base":       pd.DataFrame(index=dates_sim),
                    "ExternalPA": pd.DataFrame(index=dates_sim),
                    "ActiveExt":  pd.DataFrame(index=dates_sim),
                }

                for sim_i in range(N_SIMULATIONS):
                    r_beta = sims[sim_i, :, 0]
                    r_H    = sims[sim_i, :, 1]
                    r_E    = sims[sim_i, :, 2]
                    r_M    = sims[sim_i, :, 3]

                    # Base (internal)
                    R_base = (r_beta - f_int_matrix[sim_i]) * Z_frac + r_H * Z_frac
                    results["Base"][sim_i] = np.prod(1 + R_base) - 1

                    # External PA
                    R_extpa = (r_beta - f_ext_pa_matrix[sim_i]) * E_pct + r_M * E_pct
                    results["ExternalPA"][sim_i] = np.prod(1 + R_extpa) - 1

                    # Active Extension
                    R_actext = (r_beta - f_act_ext_matrix[sim_i]) * A_pct + r_E * A_pct
                    results["ActiveExt"][sim_i] = np.prod(1 + R_actext) - 1

                    if sim_i == 0:
                        raw_returns["Base"]       = pd.DataFrame({"Base": R_base}, index=dates_sim)
                        raw_returns["ExternalPA"] = pd.DataFrame({"ExternalPA": R_extpa}, index=dates_sim)
                        raw_returns["ActiveExt"]  = pd.DataFrame({"ActiveExt": R_actext}, index=dates_sim)

                df_yearly = pd.DataFrame(results)
                summary_rows = []
                for cfg, arr in df_yearly.items():
                    ann_ret = np.mean(arr)
                    ann_vol = np.std(arr, ddof=1)
                    var_95  = np.percentile(arr, 5)
                    te      = np.nan

                    mr_series = raw_returns[cfg].iloc[:, 0]
                    threshold = -buffer_multiple * idx_sigma
                    breach_pct = np.mean(mr_series < threshold) * 100

                    summary_rows.append({
                        "Config":            cfg,
                        "Ext %":             ext_pct * 100.0,
                        "Act %":             act_pct * 100.0,
                        "Internal PA (mm)":  Z_mm,
                        "Annual Return":     ann_ret,
                        "Annual Vol":        ann_vol,
                        "VaR 95":            var_95,
                        "TE (est.)":         te,
                        "Breach %":          breach_pct
                    })

                summary_df = pd.DataFrame(summary_rows)
                all_summaries.append(summary_df)

                for key, df_r in raw_returns.items():
                    sheet_key = f"{key}_E{int(ext_pct*100)}_A{int(act_pct*100)}"
                    all_raw_returns[sheet_key] = df_r

        final_summary = pd.concat(all_summaries, ignore_index=True)

        inputs_dict = {
            "Analysis mode":               "capital",
            "Total fund capital (mm)":     total_fund_capital,
            "Max external combined (%)":   get_num(raw_params, "max_external_combined_percent", ""),
            "External step size (%)":      get_num(raw_params, "external_step_size_percent", ""),
            "Annual financing mean (%)":   financing_mean_annual,
            "Annual financing vol (%)":    financing_vol_annual,
            "Monthly spike probability":   spike_prob,
            "Spike size (σ × multiplier)": spike_factor,
            "Buffer multiple":             buffer_multiple
        }

        export_to_excel(inputs_dict, final_summary, all_raw_returns)

        # Convert numeric summary columns to human-friendly format
        display_df = final_summary.copy()
        display_df = display_df.rename(columns={
            "Annual Return": "Annual Return (%)",
            "Annual Vol":    "Annual Volatility (%)",
            "VaR 95":        "95%-VaR (%)",
            "TE (est.)":     "Tracking Error (%)",
            "Breach %":      "Breach Probability (%)"
        })
        # Format percentages
        for col in ["Annual Return (%)", "Annual Volatility (%)", "95%-VaR (%)", "Tracking Error (%)", "Breach Probability (%)"]:
            display_df[col] = display_df[col].map("{:.1f}%".format)

        pd.set_option("display.max_rows", None)
        print("\n=== Summary Table (Capital Mode) ===\n")
        print(display_df.to_string(index=False))
        print()

        # ─── PLOTTING DISPATCH ────────────────────────────────────────────────────
        # Helper to pivot decimals to percentages for axes
        def perc(x): return x * 100.0

        if plot_heatmap_flag:
            # Heatmap of Annual Return vs Ext % & Act %
            pivot = final_summary.pivot(index="Act %", columns="Ext %", values="Annual Return")
            X = np.array(sorted(final_summary["Ext %"].unique()))
            Y = np.array(sorted(final_summary["Act %"].unique()))
            Z = pivot.values
            fig, ax = plt.subplots(figsize=(6,5))
            hm = ax.imshow(
                Z,
                origin="lower",
                aspect="auto",
                cmap="viridis",
                extent=[X.min(), X.max(), Y.min(), Y.max()]
            )
            ax.set_xlabel("Ext %")
            ax.set_ylabel("Act %")
            ax.set_title("Heatmap: Annual Return vs Ext% & Act%")
            fig.colorbar(hm, ax=ax, label="Annual Return (decimal)")
            plt.tight_layout()
            plt.show()

        if plot_line_flag:
            # Line plot: Annual Return vs Ext % for each Act % slice
            fig, ax = plt.subplots(figsize=(6,4))
            for act_level in sorted(final_summary["Act %"].unique()):
                sub = final_summary[final_summary["Act %"] == act_level]
                x_vals = sub["Ext %"].values
                y_vals = sub["Annual Return"].values * 100.0
                ax.plot(x_vals, y_vals, marker="o", label=f"Act % = {act_level:.1f}")
            ax.set_xlabel("Ext %")
            ax.set_ylabel("Annual Return (%)")
            ax.set_title("Line: Return vs Ext % by Active %")
            ax.legend()
            plt.tight_layout()
            plt.show()

        if plot_boxplot_flag:
            # Boxplot of first-simulation Monthly Returns vs Ext % at Act% = 0%
            sim_list = []
            for sheet_key, df_r in all_raw_returns.items():
                parts = sheet_key.split("_")
                cfg   = parts[0]
                try:
                    ext_pct = float(parts[1].replace("E","")) / 100.0
                    act_pct = float(parts[2].replace("A","")) / 100.0
                except:
                    continue
                for val in df_r.iloc[:, 0].values:
                    sim_list.append({
                        "Config":       cfg,
                        "Ext %":        ext_pct * 100.0,
                        "Act %":        act_pct * 100.0,
                        "Monthly return": val * 100.0
                    })
            sim_df = pd.DataFrame(sim_list)
            subset = sim_df[sim_df["Act %"] == 0.0]
            groups = [g["Monthly return"].values for _, g in subset.groupby("Ext %")]
            labels = sorted(subset["Ext %"].unique())
            fig, ax = plt.subplots(figsize=(6,4))
            ax.boxplot(groups, labels=[f"{l:.0f}" for l in labels])
            ax.set_xlabel("Ext %")
            ax.set_ylabel("Monthly Return (%)")
            ax.set_title("Boxplot: Monthly Return by Ext % (Act%=0%)")
            plt.tight_layout()
            plt.show()

        if plot_scatter_flag:
            fig, ax = plt.subplots(figsize=(5,4))
            te_vals     = final_summary["Tracking Error (%)"].astype(float) * 100.0
            breach_vals = final_summary["Breach %"].astype(float)
            ax.scatter(te_vals, breach_vals, alpha=0.6)
            ax.set_xlabel("Tracking Error (%)")
            ax.set_ylabel("Breach Probability (%)")
            ax.set_title("Scatter: TE vs Breach Probability")
            plt.tight_layout()
            plt.show()

        if plot_time_series_flag:
            threshold = -buffer_multiple * idx_sigma
            for sample_key in all_raw_returns.keys():
                if sample_key.startswith("Base_"):
                    df_path = all_raw_returns[sample_key]
                    fig, ax = plt.subplots(figsize=(6,3))
                    ax.plot(df_path.index, df_path.iloc[:, 0], label=sample_key)
                    ax.axhline(y=threshold, color="red", linestyle="--", label="Threshold")
                    ax.set_title(f"Time Series (first-sim) for {sample_key}")
                    ax.set_xlabel("Month")
                    ax.set_ylabel("Monthly return")
                    ax.legend()
                    plt.tight_layout()
                    plt.show()
                    break  # only show one example

        if plot_histogram_flag:
            # Pick a representative ExternalPA_E25_A25 if exists
            target = None
            for key in all_raw_returns:
                if key.startswith("ExternalPA_E25_A25"):
                    target = key
                    break
            if target:
                data = all_raw_returns[target].iloc[:, 0].values * 100.0
                fig, ax = plt.subplots(figsize=(5,4))
                ax.hist(data, bins=30, alpha=0.7)
                ax.set_xlabel("Monthly Return (%)")
                ax.set_ylabel("Frequency")
                ax.set_title(f"Histogram: {target}")
                plt.tight_layout()
                plt.show()

        if plot_surface_flag:
            xv = sorted(final_summary["Ext %"].unique())
            yv = sorted(final_summary["Act %"].unique())
            Z = final_summary.pivot(index="Act %", columns="Ext %", values="Breach %").values
            X, Y = np.meshgrid(np.array(xv), np.array(yv))
            fig = plt.figure(figsize=(6,5))
            ax3 = fig.add_subplot(111, projection="3d")
            surf = ax3.plot_surface(X, Y, Z, cmap="viridis", edgecolor="none")
            ax3.set_xlabel("Ext %")
            ax3.set_ylabel("Act %")
            ax3.set_zlabel("Breach %")
            ax3.set_title("Surface: Breach vs Ext & Act")
            fig.colorbar(surf, shrink=0.5, aspect=5, label="Breach %")
            plt.tight_layout()
            plt.show()

    elif analysis_mode == "returns":
        # ---------------------------------------------------------------------------------------
        # 8.9.b) Returns mode: Sweep over μ_H, σ_H, μ_E, σ_E, μ_M, σ_M. Other params fixed at midpoint.
        # ---------------------------------------------------------------------------------------

        E_mm = get_num(raw_params, "external_pa_capital", None)
        A_mm = get_num(raw_params, "active_ext_capital",  None)
        Z_mm = get_num(raw_params, "internal_pa_capital", None)
        if any(v is None for v in [E_mm, A_mm, Z_mm]):
            raise RuntimeError(
                "In 'returns' mode, please supply: External PA capital (mm), "
                "Active Extension capital (mm), Internal PA capital (mm)."
            )

        mu_H_list    = build_range("mu_H",    default_midpoint=default_mu_H)
        sigma_H_list = build_range("sigma_H", default_midpoint=default_sigma_H)
        mu_E_list    = build_range("mu_E",    default_midpoint=default_mu_E)
        sigma_E_list = build_range("sigma_E", default_midpoint=default_sigma_E)
        mu_M_list    = build_range("mu_M",    default_midpoint=default_mu_M)
        sigma_M_list = build_range("sigma_M", default_midpoint=default_sigma_M)

        ext_alpha_list = [default_ext_alpha_frac]
        act_share_list = [default_act_share]
        sd_list        = [default_sd_mult]

        all_summaries = []
        all_raw_returns = {}

        for muH in mu_H_list:
            for sH in sigma_H_list:
                for muE in mu_E_list:
                    for sE in sigma_E_list:
                        for muM_ in mu_M_list:
                            for sM in sigma_M_list:
                                f_internal = simulate_financing(
                                    12,
                                    internal_financing_mean_annual/12,
                                    internal_financing_vol_annual/12,
                                    internal_spike_prob,
                                    internal_spike_factor
                                )
                                f_ext_pa = simulate_financing(
                                    12,
                                    ext_pa_financing_mean_annual/12,
                                    ext_pa_financing_vol_annual/12,
                                    ext_pa_spike_prob,
                                    ext_pa_spike_factor
                                )
                                f_act_ext = simulate_financing(
                                    12,
                                    act_ext_financing_mean_annual/12,
                                    act_ext_financing_vol_annual/12,
                                    act_ext_spike_prob,
                                    act_ext_spike_factor
                                )

                                cov_mat = build_cov_matrix(
                                    rho_idx_H, rho_idx_E, rho_idx_M,
                                    rho_H_E, rho_H_M, rho_E_M,
                                    idx_sigma, sH, sE, sM
                                )

                                N_SIMULATIONS = get_num(raw_params, "N_SIMULATIONS", 5000)
                                N_MONTHS      = get_num(raw_params, "N_MONTHS",    12)

                                sims = np.random.multivariate_normal(
                                    [mu_idx, muH, muE, muM_],
                                    cov_mat,
                                    size=(N_SIMULATIONS, N_MONTHS)
                                )

                                f_int_matrix    = np.tile(f_internal, (N_SIMULATIONS, 1))
                                f_ext_pa_matrix = np.tile(f_ext_pa,    (N_SIMULATIONS, 1))
                                f_act_ext_matrix= np.tile(f_act_ext,   (N_SIMULATIONS, 1))

                                results = {
                                    "Base":       np.zeros(N_SIMULATIONS),
                                    "ExternalPA": np.zeros(N_SIMULATIONS),
                                    "ActiveExt":  np.zeros(N_SIMULATIONS)
                                }
                                dates_sim = pd.date_range(
                                    start=idx_series.index[-1] + pd.DateOffset(months=1),
                                    periods=N_MONTHS, freq="ME"
                                )
                                raw_returns = {
                                    "Base":       pd.DataFrame(index=dates_sim),
                                    "ExternalPA": pd.DataFrame(index=dates_sim),
                                    "ActiveExt":  pd.DataFrame(index=dates_sim),
                                }

                                for sim_i in range(N_SIMULATIONS):
                                    r_beta = sims[sim_i, :, 0]
                                    r_H    = sims[sim_i, :, 1]
                                    r_E    = sims[sim_i, :, 2]
                                    r_M    = sims[sim_i, :, 3]

                                    R_base   = (r_beta - f_int_matrix[sim_i]) + r_H
                                    R_extpa  = (r_beta - f_ext_pa_matrix[sim_i]) + r_M
                                    R_actext = (r_beta - f_act_ext_matrix[sim_i]) + r_E

                                    results["Base"][sim_i]       = np.prod(1 + R_base) - 1
                                    results["ExternalPA"][sim_i] = np.prod(1 + R_extpa) - 1
                                    results["ActiveExt"][sim_i]  = np.prod(1 + R_actext) - 1

                                    if sim_i == 0:
                                        raw_returns["Base"]       = pd.DataFrame({"Base": R_base}, index=dates_sim)
                                        raw_returns["ExternalPA"] = pd.DataFrame({"ExternalPA": R_extpa}, index=dates_sim)
                                        raw_returns["ActiveExt"]  = pd.DataFrame({"ActiveExt": R_actext}, index=dates_sim)

                                df_yearly = pd.DataFrame(results)
                                summary_rows = []
                                for cfg, arr in df_yearly.items():
                                    ann_ret = np.mean(arr)
                                    ann_vol = np.std(arr, ddof=1)
                                    var_95  = np.percentile(arr, 5)
                                    te      = np.nan

                                    mr_series = raw_returns[cfg].iloc[:, 0]
                                    threshold = -buffer_multiple * idx_sigma
                                    breach_pct = np.mean(mr_series < threshold) * 100

                                    summary_rows.append({
                                        "Config":            cfg,
                                        "μ_H (%)":           muH * 100.0,
                                        "σ_H (%)":           sH * 100.0,
                                        "μ_E (%)":           muE * 100.0,
                                        "σ_E (%)":           sE * 100.0,
                                        "μ_M (%)":           muM_ * 100.0,
                                        "σ_M (%)":           sM * 100.0,
                                        "Internal PA (mm)":  Z_mm,
                                        "External PA (mm)":  E_mm,
                                        "Active Ext (mm)":   A_mm,
                                        "Annual Return":     ann_ret,
                                        "Annual Vol":        ann_vol,
                                        "VaR 95":            var_95,
                                        "TE (est.)":         te,
                                        "Breach %":          breach_pct
                                    })

                                summary_df = pd.DataFrame(summary_rows)
                                all_summaries.append(summary_df)

                                for key, df_r in raw_returns.items():
                                    sheet_key = (
                                        f"{key}_H{int(muH*100)}"
                                        f"_sH{int(sH*100)}"
                                        f"_E{int(muE*100)}"
                                        f"_sE{int(sE*100)}"
                                        f"_M{int(muM_*100)}"
                                        f"_sM{int(sM*100)}"
                                    )
                                    all_raw_returns[sheet_key] = df_r

        final_summary = pd.concat(all_summaries, ignore_index=True)

        inputs_dict = {
            "Analysis mode":                    "returns",
            "In-House return min (%)":          get_num(raw_params, "mu_H_min",    ""),
            "In-House return max (%)":          get_num(raw_params, "mu_H_max",    ""),
            "In-House return step (%)":         get_num(raw_params, "mu_H_step",   ""),
            "In-House vol min (%)":             get_num(raw_params, "sigma_H_min", ""),
            "In-House vol max (%)":             get_num(raw_params, "sigma_H_max", ""),
            "In-House vol step (%)":            get_num(raw_params, "sigma_H_step",""),
            "Alpha-Extension return min (%)":   get_num(raw_params, "mu_E_min",    ""),
            "Alpha-Extension return max (%)":   get_num(raw_params, "mu_E_max",    ""),
            "Alpha-Extension return step (%)":  get_num(raw_params, "mu_E_step",   ""),
            "Alpha-Extension vol min (%)":      get_num(raw_params, "sigma_E_min", ""),
            "Alpha-Extension vol max (%)":      get_num(raw_params, "sigma_E_max", ""),
            "Alpha-Extension vol step (%)":     get_num(raw_params, "sigma_E_step",""),
            "External return min (%)":          get_num(raw_params, "mu_M_min",    ""),
            "External return max (%)":          get_num(raw_params, "mu_M_max",    ""),
            "External return step (%)":         get_num(raw_params, "mu_M_step",   ""),
            "External vol min (%)":             get_num(raw_params, "sigma_M_min", ""),
            "External vol max (%)":             get_num(raw_params, "sigma_M_max", ""),
            "External vol step (%)":            get_num(raw_params, "sigma_M_step",""),
            "External PA capital (mm)":         E_mm,
            "Active Extension capital (mm)":    A_mm,
            "Internal PA capital (mm)":         Z_mm,
            "Buffer multiple":                  buffer_multiple
        }

        export_to_excel(inputs_dict, final_summary, all_raw_returns)

        display_df = final_summary.copy()
        display_df = display_df.rename(columns={
            "Annual Return": "Annual Return (%)",
            "Annual Vol":    "Annual Volatility (%)",
            "VaR 95":        "95%-VaR (%)",
            "TE (est.)":     "Tracking Error (%)",
            "Breach %":      "Breach Probability (%)"
        })
        for col in ["Annual Return (%)", "Annual Volatility (%)", "95%-VaR (%)", "Tracking Error (%)", "Breach Probability (%)"]:
            display_df[col] = display_df[col].map("{:.1f}%".format)

        pd.set_option("display.max_rows", None)
        print("\n=== Summary Table (Returns Mode) ===\n")
        print(display_df.to_string(index=False))
        print()

        # ─── PLOTTING DISPATCH ────────────────────────────────────────────────────

        if plot_heatmap_flag:
            # Heatmap: pick two parameters to visualize, e.g. μ_H vs σ_H → Annual Return
            pivot = final_summary.pivot(index="σ_H (%)", columns="μ_H (%)", values="Annual Return")
            X = np.array(sorted(final_summary["μ_H (%)"].unique()))
            Y = np.array(sorted(final_summary["σ_H (%)"].unique()))
            Z = pivot.values
            fig, ax = plt.subplots(figsize=(6,5))
            hm = ax.imshow(
                Z,
                origin="lower",
                aspect="auto",
                cmap="viridis",
                extent=[X.min(), X.max(), Y.min(), Y.max()]
            )
            ax.set_xlabel("μ_H (%)")
            ax.set_ylabel("σ_H (%)")
            ax.set_title("Heatmap: Annual Return vs μ_H & σ_H")
            fig.colorbar(hm, ax=ax, label="Annual Return (decimal)")
            plt.tight_layout()
            plt.show()

        if plot_line_flag:
            # Line: sweep μ_H for fixed σ_H (take first σ_H level)
            first_sigma = sorted(final_summary["σ_H (%)"].unique())[0]
            sub = final_summary[final_summary["σ_H (%)"] == first_sigma]
            x_vals = sub["μ_H (%)"].values
            y_vals = sub["Annual Return"].values * 100.0
            fig, ax = plt.subplots(figsize=(6,4))
            ax.plot(x_vals, y_vals, marker="o")
            ax.set_xlabel("μ_H (%)")
            ax.set_ylabel("Annual Return (%)")
            ax.set_title(f"Line: μ_H vs Return (σ_H={first_sigma:.1f}%)")
            plt.tight_layout()
            plt.show()

        if plot_boxplot_flag:
            # Boxplot of monthly returns for one parameter combination, e.g. first in grid
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.boxplot(data, labels=[sample_key])
            ax.set_ylabel("Monthly Return (%)")
            ax.set_title(f"Boxplot: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_scatter_flag:
            fig, ax = plt.subplots(figsize=(5,4))
            te_vals     = final_summary["Tracking Error (%)"].astype(float) * 100.0
            breach_vals = final_summary["Breach %"].astype(float)
            ax.scatter(te_vals, breach_vals, alpha=0.6)
            ax.set_xlabel("Tracking Error (%)")
            ax.set_ylabel("Breach Probability (%)")
            ax.set_title("Scatter: TE vs Breach Probability")
            plt.tight_layout()
            plt.show()

        if plot_time_series_flag:
            threshold = -buffer_multiple * idx_sigma
            sample_key = list(all_raw_returns.keys())[0]
            df_path = all_raw_returns[sample_key]
            fig, ax = plt.subplots(figsize=(6,3))
            ax.plot(df_path.index, df_path.iloc[:, 0], label=sample_key)
            ax.axhline(y=threshold, color="red", linestyle="--", label="Threshold")
            ax.set_title(f"Time Series (first-sim) for {sample_key}")
            ax.set_xlabel("Month")
            ax.set_ylabel("Monthly return")
            ax.legend()
            plt.tight_layout()
            plt.show()

        if plot_histogram_flag:
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.hist(data, bins=30, alpha=0.7)
            ax.set_xlabel("Monthly Return (%)")
            ax.set_ylabel("Frequency")
            ax.set_title(f"Histogram: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_surface_flag:
            xv = sorted(final_summary["μ_H (%)"].unique())
            yv = sorted(final_summary["σ_H (%)"].unique())
            Z = final_summary.pivot(index="σ_H (%)", columns="μ_H (%)", values="Breach %").values
            X, Y = np.meshgrid(np.array(xv), np.array(yv))
            fig = plt.figure(figsize=(6,5))
            ax3 = fig.add_subplot(111, projection="3d")
            surf = ax3.plot_surface(X, Y, Z, cmap="viridis", edgecolor="none")
            ax3.set_xlabel("μ_H (%)")
            ax3.set_ylabel("σ_H (%)")
            ax3.set_zlabel("Breach %")
            ax3.set_title("Surface: Breach vs μ_H & σ_H")
            fig.colorbar(surf, shrink=0.5, aspect=5, label="Breach %")
            plt.tight_layout()
            plt.show()

    elif analysis_mode == "alpha_shares":
        # ---------------------------------------------------------------------------------------
        # 8.9.c) Alpha_Shares mode: Sweep over external_pa_alpha_frac, active_share. Others fixed.
        # ---------------------------------------------------------------------------------------

        E_mm = get_num(raw_params, "external_pa_capital", None)
        A_mm = get_num(raw_params, "active_ext_capital",  None)
        Z_mm = get_num(raw_params, "internal_pa_capital", None)
        if any(v is None for v in [E_mm, A_mm, Z_mm]):
            raise RuntimeError(
                "In 'alpha_shares' mode, supply: External PA capital (mm), "
                "Active Extension capital (mm), Internal PA capital (mm)."
            )

        external_pa_alpha_frac_list = build_range("external_pa_alpha_frac", default_midpoint=default_ext_alpha_frac)
        active_share_list           = build_range("active_share",            default_midpoint=default_act_share)

        mu_H_list    = [default_mu_H]
        sigma_H_list = [default_sigma_H]
        mu_E_list    = [default_mu_E]
        sigma_E_list = [default_sigma_E]
        mu_M_list    = [default_mu_M]
        sigma_M_list = [default_sigma_M]
        sd_list      = [default_sd_mult]

        all_summaries = []
        all_raw_returns = {}

        for ext_alpha in external_pa_alpha_frac_list:
            for act_share in active_share_list:
                f_internal = simulate_financing(
                    12,
                    internal_financing_mean_annual/12,
                    internal_financing_vol_annual/12,
                    internal_spike_prob,
                    internal_spike_factor
                )
                f_ext_pa = simulate_financing(
                    12,
                    ext_pa_financing_mean_annual/12,
                    ext_pa_financing_vol_annual/12,
                    ext_pa_spike_prob,
                    ext_pa_spike_factor
                )
                f_act_ext = simulate_financing(
                    12,
                    act_ext_financing_mean_annual/12,
                    act_ext_financing_vol_annual/12,
                    act_ext_spike_prob,
                    act_ext_spike_factor
                )

                cov_mat = build_cov_matrix(
                    rho_idx_H, rho_idx_E, rho_idx_M,
                    rho_H_E, rho_H_M, rho_E_M,
                    idx_sigma, default_sigma_H, default_sigma_E, default_sigma_M
                )

                N_SIMULATIONS = get_num(raw_params, "N_SIMULATIONS", 5000)
                N_MONTHS      = get_num(raw_params, "N_MONTHS",    12)

                sims = np.random.multivariate_normal(
                    [mu_idx, default_mu_H, default_mu_E, default_mu_M],
                    cov_mat,
                    size=(N_SIMULATIONS, N_MONTHS)
                )

                f_int_matrix    = np.tile(f_internal, (N_SIMULATIONS, 1))
                f_ext_pa_matrix = np.tile(f_ext_pa,    (N_SIMULATIONS, 1))
                f_act_ext_matrix= np.tile(f_act_ext,   (N_SIMULATIONS, 1))

                results = {
                    "Base":       np.zeros(N_SIMULATIONS),
                    "ExternalPA": np.zeros(N_SIMULATIONS),
                    "ActiveExt":  np.zeros(N_SIMULATIONS)
                }
                dates_sim = pd.date_range(
                    start=idx_series.index[-1] + pd.DateOffset(months=1),
                    periods=N_MONTHS, freq="ME"
                )
                raw_returns = {
                    "Base":       pd.DataFrame(index=dates_sim),
                    "ExternalPA": pd.DataFrame(index=dates_sim),
                    "ActiveExt":  pd.DataFrame(index=dates_sim),
                }

                for sim_i in range(N_SIMULATIONS):
                    r_beta = sims[sim_i, :, 0]
                    r_H    = sims[sim_i, :, 1]
                    r_E    = sims[sim_i, :, 2]
                    r_M    = sims[sim_i, :, 3]

                    R_base   = (r_beta - f_int_matrix[sim_i]) + r_H
                    R_extpa  = (r_beta - f_ext_pa_matrix[sim_i]) * ext_alpha + r_M * ext_alpha
                    R_actext = (r_beta - f_act_ext_matrix[sim_i]) * act_share + r_E * act_share

                    results["Base"][sim_i]       = np.prod(1 + R_base) - 1
                    results["ExternalPA"][sim_i] = np.prod(1 + R_extpa) - 1
                    results["ActiveExt"][sim_i]  = np.prod(1 + R_actext) - 1

                    if sim_i == 0:
                        raw_returns["Base"]       = pd.DataFrame({"Base": R_base}, index=dates_sim)
                        raw_returns["ExternalPA"] = pd.DataFrame({"ExternalPA": R_extpa}, index=dates_sim)
                        raw_returns["ActiveExt"]  = pd.DataFrame({"ActiveExt": R_actext}, index=dates_sim)

                df_yearly = pd.DataFrame(results)
                summary_rows = []
                for cfg, arr in df_yearly.items():
                    ann_ret = np.mean(arr)
                    ann_vol = np.std(arr, ddof=1)
                    var_95  = np.percentile(arr, 5)
                    te      = np.nan

                    mr_series = raw_returns[cfg].iloc[:, 0]
                    threshold = -buffer_multiple * idx_sigma
                    breach_pct = np.mean(mr_series < threshold) * 100

                    summary_rows.append({
                        "Config":            cfg,
                        "External PA α (%)": ext_alpha * 100.0,
                        "Active share (%)":  act_share * 100.0,
                        "Internal PA (mm)":  Z_mm,
                        "External PA (mm)":  E_mm,
                        "Active Ext (mm)":   A_mm,
                        "Annual Return":     ann_ret,
                        "Annual Vol":        ann_vol,
                        "VaR 95":            var_95,
                        "TE (est.)":         te,
                        "Breach %":          breach_pct
                    })

                summary_df = pd.DataFrame(summary_rows)
                all_summaries.append(summary_df)

                for key, df_r in raw_returns.items():
                    sheet_key = f"{key}_α{int(ext_alpha*100)}_act{int(act_share*100)}"
                    all_raw_returns[sheet_key] = df_r

        final_summary = pd.concat(all_summaries, ignore_index=True)

        inputs_dict = {
            "Analysis mode":                  "alpha_shares",
            "External PA α fraction min (%)": get_num(raw_params, "external_pa_alpha_frac_min", ""),
            "External PA α fraction max (%)": get_num(raw_params, "external_pa_alpha_frac_max", ""),
            "External PA α fraction step (%)":get_num(raw_params, "external_pa_alpha_frac_step", ""),
            "Active share min (%)":          get_num(raw_params, "active_share_min", ""),
            "Active share max (%)":          get_num(raw_params, "active_share_max", ""),
            "Active share step (%)":         get_num(raw_params, "active_share_step", ""),
            "External PA capital (mm)":      E_mm,
            "Active Extension capital (mm)": A_mm,
            "Internal PA capital (mm)":      Z_mm,
            "Buffer multiple":               buffer_multiple
        }

        export_to_excel(inputs_dict, final_summary, all_raw_returns)

        display_df = final_summary.copy()
        display_df = display_df.rename(columns={
            "Annual Return": "Annual Return (%)",
            "Annual Vol":    "Annual Volatility (%)",
            "VaR 95":        "95%-VaR (%)",
            "TE (est.)":     "Tracking Error (%)",
            "Breach %":      "Breach Probability (%)"
        })
        for col in ["Annual Return (%)", "Annual Volatility (%)", "95%-VaR (%)", "Tracking Error (%)", "Breach Probability (%)"]:
            display_df[col] = display_df[col].map("{:.1f}%".format)

        pd.set_option("display.max_rows", None)
        print("\n=== Summary Table (Alpha_Shares Mode) ===\n")
        print(display_df.to_string(index=False))
        print()

        # ─── PLOTTING DISPATCH ────────────────────────────────────────────────────

        if plot_heatmap_flag:
            pivot = final_summary.pivot(index="Active share (%)", columns="External PA α (%)", values="Annual Return")
            X = np.array(sorted(final_summary["External PA α (%)"].unique()))
            Y = np.array(sorted(final_summary["Active share (%)"].unique()))
            Z = pivot.values
            fig, ax = plt.subplots(figsize=(6,5))
            hm = ax.imshow(
                Z,
                origin="lower",
                aspect="auto",
                cmap="viridis",
                extent=[X.min(), X.max(), Y.min(), Y.max()]
            )
            ax.set_xlabel("External PA α (%)")
            ax.set_ylabel("Active share (%)")
            ax.set_title("Heatmap: Annual Return vs α & Active share")
            fig.colorbar(hm, ax=ax, label="Annual Return (decimal)")
            plt.tight_layout()
            plt.show()

        if plot_line_flag:
            first_act = sorted(final_summary["Active share (%)"].unique())[0]
            sub = final_summary[final_summary["Active share (%)"] == first_act]
            x_vals = sub["External PA α (%)"].values
            y_vals = sub["Annual Return"].values * 100.0
            fig, ax = plt.subplots(figsize=(6,4))
            ax.plot(x_vals, y_vals, marker="o")
            ax.set_xlabel("External PA α (%)")
            ax.set_ylabel("Annual Return (%)")
            ax.set_title(f"Line: α vs Return (Active share={first_act:.1f}%)")
            plt.tight_layout()
            plt.show()

        if plot_boxplot_flag:
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.boxplot(data, labels=[sample_key])
            ax.set_ylabel("Monthly Return (%)")
            ax.set_title(f"Boxplot: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_scatter_flag:
            fig, ax = plt.subplots(figsize=(5,4))
            te_vals     = final_summary["Tracking Error (%)"].astype(float) * 100.0
            breach_vals = final_summary["Breach %"].astype(float)
            ax.scatter(te_vals, breach_vals, alpha=0.6)
            ax.set_xlabel("Tracking Error (%)")
            ax.set_ylabel("Breach Probability (%)")
            ax.set_title("Scatter: TE vs Breach Probability")
            plt.tight_layout()
            plt.show()

        if plot_time_series_flag:
            threshold = -buffer_multiple * idx_sigma
            sample_key = list(all_raw_returns.keys())[0]
            df_path = all_raw_returns[sample_key]
            fig, ax = plt.subplots(figsize=(6,3))
            ax.plot(df_path.index, df_path.iloc[:, 0], label=sample_key)
            ax.axhline(y=threshold, color="red", linestyle="--", label="Threshold")
            ax.set_title(f"Time Series (first-sim) for {sample_key}")
            ax.set_xlabel("Month")
            ax.set_ylabel("Monthly return")
            ax.legend()
            plt.tight_layout()
            plt.show()

        if plot_histogram_flag:
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.hist(data, bins=30, alpha=0.7)
            ax.set_xlabel("Monthly Return (%)")
            ax.set_ylabel("Frequency")
            ax.set_title(f"Histogram: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_surface_flag:
            xv = sorted(final_summary["External PA α (%)"].unique())
            yv = sorted(final_summary["Active share (%)"].unique())
            Z = final_summary.pivot(index="Active share (%)", columns="External PA α (%)", values="Breach %").values
            X, Y = np.meshgrid(np.array(xv), np.array(yv))
            fig = plt.figure(figsize=(6,5))
            ax3 = fig.add_subplot(111, projection="3d")
            surf = ax3.plot_surface(X, Y, Z, cmap="viridis", edgecolor="none")
            ax3.set_xlabel("External PA α (%)")
            ax3.set_ylabel("Active share (%)")
            ax3.set_zlabel("Breach %")
            ax3.set_title("Surface: Breach vs α & Active share")
            fig.colorbar(surf, shrink=0.5, aspect=5, label="Breach %")
            plt.tight_layout()
            plt.show()

    elif analysis_mode == "vol_mult":
        # ---------------------------------------------------------------------------------------
        # 8.9.d) Vol_Mult mode: Sweep SD multiple; other params fixed at midpoint.
        # ---------------------------------------------------------------------------------------

        E_mm = get_num(raw_params, "external_pa_capital", None)
        A_mm = get_num(raw_params, "active_ext_capital",  None)
        Z_mm = get_num(raw_params, "internal_pa_capital", None)
        if any(v is None for v in [E_mm, A_mm, Z_mm]):
            raise RuntimeError(
                "In 'vol_mult' mode, supply: External PA capital (mm), "
                "Active Extension capital (mm), Internal PA capital (mm)."
            )

        sd_list = build_range_int("sd_of_vol_mult", default_midpoint=default_sd_mult)

        mu_H_list    = [default_mu_H]
        sigma_H_list = [default_sigma_H]
        mu_E_list    = [default_mu_E]
        sigma_E_list = [default_sigma_E]
        mu_M_list    = [default_mu_M]
        sigma_M_list = [default_sigma_M]
        ext_alpha_list = [default_ext_alpha_frac]
        act_share_list = [default_act_share]

        all_summaries = []
        all_raw_returns = {}

        for sd_mult in sd_list:
            f_internal = simulate_financing(
                12,
                internal_financing_mean_annual/12,
                internal_financing_vol_annual/12,
                internal_spike_prob,
                internal_spike_factor
            )
            f_ext_pa = simulate_financing(
                12,
                ext_pa_financing_mean_annual/12,
                ext_pa_financing_vol_annual/12,
                ext_pa_spike_prob,
                ext_pa_spike_factor
            )
            f_act_ext = simulate_financing(
                12,
                act_ext_financing_mean_annual/12,
                act_ext_financing_vol_annual/12,
                act_ext_spike_prob,
                act_ext_spike_factor
            )

            cov_mat = build_cov_matrix(
                rho_idx_H, rho_idx_E, rho_idx_M,
                rho_H_E, rho_H_M, rho_E_M,
                idx_sigma, default_sigma_H, default_sigma_E, default_sigma_M
            )

            N_SIMULATIONS = get_num(raw_params, "N_SIMULATIONS", 5000)
            N_MONTHS      = get_num(raw_params, "N_MONTHS",    12)

            sims = np.random.multivariate_normal(
                [mu_idx, default_mu_H, default_mu_E, default_mu_M],
                cov_mat,
                size=(N_SIMULATIONS, N_MONTHS)
            )

            f_int_matrix    = np.tile(f_internal, (N_SIMULATIONS, 1))
            f_ext_pa_matrix = np.tile(f_ext_pa,    (N_SIMULATIONS, 1))
            f_act_ext_matrix= np.tile(f_act_ext,   (N_SIMULATIONS, 1))

            results = {
                "Base":       np.zeros(N_SIMULATIONS),
                "ExternalPA": np.zeros(N_SIMULATIONS),
                "ActiveExt":  np.zeros(N_SIMULATIONS)
            }
            dates_sim = pd.date_range(
                start=idx_series.index[-1] + pd.DateOffset(months=1),
                periods=N_MONTHS, freq="ME"
            )
            raw_returns = {
                "Base":       pd.DataFrame(index=dates_sim),
                "ExternalPA": pd.DataFrame(index=dates_sim),
                "ActiveExt":  pd.DataFrame(index=dates_sim),
            }

            for sim_i in range(N_SIMULATIONS):
                r_beta = sims[sim_i, :, 0]
                r_H    = sims[sim_i, :, 1]
                r_E    = sims[sim_i, :, 2]
                r_M    = sims[sim_i, :, 3]

                R_base   = (r_beta - f_int_matrix[sim_i]) + r_H
                R_extpa  = (r_beta - f_ext_pa_matrix[sim_i]) * default_ext_alpha_frac + r_M * default_ext_alpha_frac
                R_actext = (r_beta - f_act_ext_matrix[sim_i]) * default_act_share + r_E * default_act_share

                results["Base"][sim_i]       = np.prod(1 + R_base) - 1
                results["ExternalPA"][sim_i] = np.prod(1 + R_extpa) - 1
                results["ActiveExt"][sim_i]  = np.prod(1 + R_actext) - 1

                if sim_i == 0:
                    raw_returns["Base"]       = pd.DataFrame({"Base": R_base}, index=dates_sim)
                    raw_returns["ExternalPA"] = pd.DataFrame({"ExternalPA": R_extpa}, index=dates_sim)
                    raw_returns["ActiveExt"]  = pd.DataFrame({"ActiveExt": R_actext}, index=dates_sim)

            df_yearly = pd.DataFrame(results)
            summary_rows = []
            for cfg, arr in df_yearly.items():
                ann_ret = np.mean(arr)
                ann_vol = np.std(arr, ddof=1)
                var_95  = np.percentile(arr, 5)
                te      = np.nan

                mr_series = raw_returns[cfg].iloc[:, 0]
                threshold = - (sd_mult * idx_sigma)
                breach_pct = np.mean(mr_series < threshold) * 100

                summary_rows.append({
                    "Config":        cfg,
                    "SD mult":       sd_mult,
                    "Internal PA (mm)": Z_mm,
                    "External PA (mm)": E_mm,
                    "Active Ext (mm)":  A_mm,
                    "Annual Return":      ann_ret,
                    "Annual Vol":         ann_vol,
                    "VaR 95":             var_95,
                    "TE (est.)":          te,
                    "Breach %":           breach_pct
                })

            summary_df = pd.DataFrame(summary_rows)
            all_summaries.append(summary_df)

            for key, df_r in raw_returns.items():
                sheet_key = f"{key}_SD{sd_mult}"
                all_raw_returns[sheet_key] = df_r

        final_summary = pd.concat(all_summaries, ignore_index=True)

        inputs_dict = {
            "Analysis mode":        "vol_mult",
            "SD multiple min":      get_num(raw_params, "sd_of_vol_mult_min", ""),
            "SD multiple max":      get_num(raw_params, "sd_of_vol_mult_max", ""),
            "SD multiple step":     get_num(raw_params, "sd_of_vol_mult_step", ""),
            "External PA capital (mm)":     E_mm,
            "Active Extension capital (mm)":A_mm,
            "Internal PA capital (mm)":     Z_mm,
            "Buffer multiple":      buffer_multiple
        }

        export_to_excel(inputs_dict, final_summary, all_raw_returns)

        display_df = final_summary.copy()
        display_df = display_df.rename(columns={
            "Annual Return": "Annual Return (%)",
            "Annual Vol":    "Annual Volatility (%)",
            "VaR 95":        "95%-VaR (%)",
            "TE (est.)":     "Tracking Error (%)",
            "Breach %":      "Breach Probability (%)"
        })
        for col in ["Annual Return (%)", "Annual Volatility (%)", "95%-VaR (%)", "Tracking Error (%)", "Breach Probability (%)"]:
            display_df[col] = display_df[col].map("{:.1f}%".format)

        pd.set_option("display.max_rows", None)
        print("\n=== Summary Table (Vol_Mult Mode) ===\n")
        print(display_df.to_string(index=False))
        print()

        # ─── PLOTTING DISPATCH ────────────────────────────────────────────────────

        if plot_heatmap_flag:
            pivot = final_summary.pivot(index="SD mult", columns="Config", values="Annual Return")
            # Example: heatmap of SD mult vs Config (Base, ExternalPA, ActiveExt) → annual return
            X = np.arange(len(pivot.columns))
            Y = pivot.index.values
            Z = pivot.values
            fig, ax = plt.subplots(figsize=(6,5))
            hm = ax.imshow(
                Z,
                origin="lower",
                aspect="auto",
                cmap="viridis",
                extent=[0, len(X)-1, Y.min(), Y.max()]
            )
            ax.set_xticks(X)
            ax.set_xticklabels(pivot.columns, rotation=45, ha="right")
            ax.set_ylabel("SD mult")
            ax.set_title("Heatmap: Annual Return vs Config & SD mult")
            fig.colorbar(hm, ax=ax, label="Annual Return (decimal)")
            plt.tight_layout()
            plt.show()

        if plot_line_flag:
            # Line: For each Config, plot Annual Return vs SD mult
            fig, ax = plt.subplots(figsize=(6,4))
            for cfg in final_summary["Config"].unique():
                sub = final_summary[final_summary["Config"] == cfg]
                x_vals = sub["SD mult"].values
                y_vals = sub["Annual Return"].values * 100.0
                ax.plot(x_vals, y_vals, marker="o", label=cfg)
            ax.set_xlabel("SD mult")
            ax.set_ylabel("Annual Return (%)")
            ax.set_title("Line: Return vs SD mult by Config")
            ax.legend()
            plt.tight_layout()
            plt.show()

        if plot_boxplot_flag:
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.boxplot(data, labels=[sample_key])
            ax.set_ylabel("Monthly Return (%)")
            ax.set_title(f"Boxplot: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_scatter_flag:
            fig, ax = plt.subplots(figsize=(5,4))
            te_vals     = final_summary["Tracking Error (%)"].astype(float) * 100.0
            breach_vals = final_summary["Breach %"].astype(float)
            ax.scatter(te_vals, breach_vals, alpha=0.6)
            ax.set_xlabel("Tracking Error (%)")
            ax.set_ylabel("Breach Probability (%)")
            ax.set_title("Scatter: TE vs Breach Probability")
            plt.tight_layout()
            plt.show()

        if plot_time_series_flag:
            threshold = -buffer_multiple * idx_sigma
            sample_key = list(all_raw_returns.keys())[0]
            df_path = all_raw_returns[sample_key]
            fig, ax = plt.subplots(figsize=(6,3))
            ax.plot(df_path.index, df_path.iloc[:, 0], label=sample_key)
            ax.axhline(y=threshold, color="red", linestyle="--", label="Threshold")
            ax.set_title(f"Time Series (first-sim) for {sample_key}")
            ax.set_xlabel("Month")
            ax.set_ylabel("Monthly return")
            ax.legend()
            plt.tight_layout()
            plt.show()

        if plot_histogram_flag:
            sample_key = list(all_raw_returns.keys())[0]
            data = all_raw_returns[sample_key].iloc[:, 0].values * 100.0
            fig, ax = plt.subplots(figsize=(5,4))
            ax.hist(data, bins=30, alpha=0.7)
            ax.set_xlabel("Monthly Return (%)")
            ax.set_ylabel("Frequency")
            ax.set_title(f"Histogram: {sample_key}")
            plt.tight_layout()
            plt.show()

        if plot_surface_flag:
            xv = sorted(final_summary["SD mult"].unique())
            configs = final_summary["Config"].unique()
            Z = final_summary.pivot(index="SD mult", columns="Config", values="Breach %").values
            X, Y = np.meshgrid(np.array(range(len(configs))), np.array(xv))
            fig = plt.figure(figsize=(6,5))
            ax3 = fig.add_subplot(111, projection="3d")
            surf = ax3.plot_surface(X, Y, Z, cmap="viridis", edgecolor="none")
            ax3.set_xticks(range(len(configs)))
            ax3.set_xticklabels(configs, rotation=45, ha="right")
            ax3.set_ylabel("SD mult")
            ax3.set_zlabel("Breach %")
            ax3.set_title("Surface: Breach vs Config & SD mult")
            fig.colorbar(surf, shrink=0.5, aspect=5, label="Breach %")
            plt.tight_layout()
            plt.show()

    else:
        raise RuntimeError(f"Unexpected analysis mode: '{analysis_mode}'")


# In[1]:


import pandas as pd
from pathlib import Path
import openpyxl

def export_everything_to_excel(
    inputs_dict: dict,
    summary_df: pd.DataFrame,
    raw_returns_dict: dict,
    index_csv_path: Path,
    python_code_path: Path,
    documentation_path: Path,
    output_filename: str = "Everything.xlsx"
):
    """
    Write a single Excel workbook with multiple tabs:
      1) Inputs         → inputs_dict + a row for index_csv_path
      2) Summary        → summary_df
      3+) Raw Returns   → one sheet per key in raw_returns_dict
      n-1) Code        → full contents of python_code_path (one line/cell)
      n  ) Documentation → full contents of documentation_path (one line/cell)

    - inputs_dict:      { parameter_name: value }
    - summary_df:       pandas DataFrame of human‐friendly summary (no index)
    - raw_returns_dict: { sheet_name: DataFrame } (each DataFrame indexed by Date)
    - index_csv_path:   Path to the index CSV file (we’ll dump that path into the Inputs tab)
    - python_code_path: Path to “portable_alpha_model.py”
    - documentation_path: Path to “model_documentation.md”
    - output_filename:  Name of the final Excel file to write
    """
    # 1) Build a DataFrame for Inputs, inserting “Index CSV (full path)” at the top
    df_inputs = pd.DataFrame.from_dict(inputs_dict, orient="index", columns=["Value"])
    df_inputs.index.name = "Parameter"
    df_inputs.reset_index(inplace=True)

    # Insert the index‐CSV row at the very top
    df_inputs.loc[-1] = ["Index CSV (full path)", str(index_csv_path)]
    df_inputs.index = df_inputs.index + 1
    df_inputs.sort_index(inplace=True)

    # 2) Prepare to write everything into one Excel file
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        # --- Inputs tab ---
        df_inputs.to_excel(writer, sheet_name="Inputs", index=False)

        # --- Summary tab ---
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # --- Raw Returns tabs ---
        # Excel limits sheet names to 31 characters; truncate if needed.
        for sheet_name, df_rr in raw_returns_dict.items():
            safe_name = sheet_name if len(sheet_name) <= 31 else sheet_name[:28] + "..."
            df_rr.to_excel(writer, sheet_name=safe_name, index=True)

        # --- Code tab ---
        # Read the .py file line by line and dump into “Code” sheet, column A
        code_lines = Path(python_code_path).read_text(encoding="utf-8").splitlines()
        wb = writer.book
        ws_code = wb.create_sheet(title="Code")
        for row_idx, line in enumerate(code_lines, start=1):
            ws_code.cell(row=row_idx, column=1, value=line)
        # Widen column A so code is readable
        ws_code.column_dimensions["A"].width = 100
        ws_code.sheet_view.showGridLines = False

        # --- Documentation tab ---
        # Read the Markdown (or plain‐text) file and dump into “Documentation” sheet
        doc_lines = Path(documentation_path).read_text(encoding="utf-8").splitlines()
        ws_doc = wb.create_sheet(title="Documentation")
        for row_idx, line in enumerate(doc_lines, start=1):
            ws_doc.cell(row=row_idx, column=1, value=line)
        ws_doc.column_dimensions["A"].width = 80
        ws_doc.sheet_view.showGridLines = False

    print(f"All tabs exported to {output_filename}.")



# # Portable Alpha + Active Extension Model Specification
# 
# Below is a comprehensive description of the updated portable‐alpha + active‐extension model, ready to paste into a Markdown cell. Every section is clearly labeled, and all equations use LaTeX delimiters.
# 
# ---
# 
# ## 1. Purpose and High-Level Overview
# 
# **Goal:**  
# Construct a Monte Carlo framework that allocates a fixed pool of capital (e.g. \$1 b) across three “sleeves” (Internal, External Portable-Alpha, and Active Extension), simulates joint returns on Index, In-House α, Extension α, and External PA α, and then reports portfolio metrics (annual return, volatility, VaR, tracking error, breach probability).
# 
# Key innovations vs. a simpler portable-alpha model:  
# 1. **Separate “reference period”** used to compute index volatility σₙ, which in turn determines the cash/margin needed to synthetically hold 1:1 index exposure.  
# 2. **Three explicit buckets** whose dollar-amounts sum to \$ 1 b, avoiding any double-counting of β + α exposures.  
# 3. **Active Extension bucket** that can be “150/50” or “170/70” long/short, specified by an “Active share (%)” input. By default, we assume 150/50 (i.e. Active share = 50 %) unless the user overrides.
# 
# Everything ultimately flows into a set of formulas—one per bucket—that map monthly draws of
# \[
# (r_{\beta},\,r_{H},\,r_{E},\,r_{M}) 
# \quad\text{and}\quad
# f_t
# \]
# into portfolio returns.
# 
# ---
# 
# ## 2. Core Assumptions and Variables
# 
# 1. **Index (β) returns**  
#    - We load a historical time series of monthly total returns on the S&P 500 TR (or whichever index) from a CSV.  
#    - We partition that series into:  
#      1. A **reference window** (e.g. 2010 – 2014) used to compute “reference volatility” σₙ.  
#      2. An **analysis window** (e.g. 2015 – 2020) used to compute the actual mean (μₙ) and volatility (σₙ) that drive our Monte Carlo draws.
# 
# 2. **Three α-streams** (simulated jointly with β)  
#    - **In-House α** \($r_H$\):  
#      - Mean = μ_H/12  
#      - Vol = σ_H / √12  
#      - Correlation ρ_{β,H} with β.  
#    - **Extension α** \($r_E$\):  
#      - Mean = μ_E/12  
#      - Vol = σ_E / √12  
#      - Correlation ρ_{β,E} with β.  
#    - **External PA α** \($r_M$\):  
#      - Mean = μ_M/12  
#      - Vol = σ_M / √12  
#      - Correlation ρ_{β,M} with β.
# 
# 3. **Financing spread** \($f_t$\)  
#    - A month-by-month random draw around a drift (financing_mean/12) with vol (financing_vol/12) and occasional jumps of size (spike_factor × (financing_vol/12)), happening with probability spike_prob.  
#    - In each month, any bucket that holds \((r_{\beta} − f_t)\) is charged that financing cost.
# 
# 4. **Total fund capital** (in millions, default = 1000)  
#    - We allocate exactly \$ 1 b across three buckets (plus any residual “cash-leftover” after margin).
# 
# 5. **Standard-deviation multiple** (sd_of_vol_mult, default = 3)  
#    - “To hold \$ 1 b of index exposure, you must keep aside cash = σₙ × (sd_of_vol_mult) × \$ 1 b.”  
#    - That cash is the **internal beta-backing** or “margin cash,” needed for futures/swaps.
# 
# 6. **Three capital buckets** (all in \$ mm, must sum to 1000)  
#    1. **External PA capital** \($X$\)  
#       - Manager takes \$ X m; buys \$ X m of index (β) and \((external_pa_alpha_frac × X m)\) of α.  
#       - Default α fraction = 50 % (\(\theta_{\mathrm{ExtPA}}=0.50\)).  
#    2. **Active Extension capital** \($Y$\)  
#       - Manager runs a long/short portfolio with **Active share** \(S\).  
#       - By default, “150/50” means \(S=0.50\) (i.e. 150 % long, 50 % short → net 100 %).  
#    3. **Internal PA capital** \($Z$\)  
#       - Runs in-house α; the remainder of internal cash (beyond margin) is used here.
# 
# 7. **Internal beta backing** \($W$\) (computed, not user-entered)  
#    \[
#      W = \sigma_{\text{ref}} \times (\mathrm{sd\_of\_vol\_mult}) \times 1000 \quad (\text{\$ mm}).
#    \]
#    - That cash sits in reserve to back a \$ 1 b index position via futures/swaps.  
#    - Because the external PA and active-extension managers each hold index exposure “inside” their \$ X m or \$ Y m, **you do not hold margin for that portion**. You only hold \(W\) for the total \$ 1 b.
# 
# ---
# 
# ## 3. Capital-Allocation Equations
# 
# 1. **Check**:  
#    \[
#      X + Y + Z \;=\; 1000 \quad(\text{\$ mm}),
#    \]  
#    where  
#    - \(X = \text{external\_pa\_capital},\)  
#    - \(Y = \text{active\_ext\_capital},\)  
#    - \(Z = \text{internal\_pa\_capital}.\)
# 
# 2. **Margin (internal beta backing)**:  
#    \[
#      W = \sigma_{\text{ref}} \times (\mathrm{sd\_of\_vol\_mult}) \times 1000 \quad (\text{\$ mm}).
#    \]
# 
# 3. **Internal cash leftover (runs In-House PA)**:  
#    \[
#      \text{internal\_cash\_leftover} 
#      = 1000 - W - Z \quad (\text{\$ mm}).
#    \]
# 
#    - If \(W + Z > 1000\), the capital structure is infeasible (you cannot hold margin + in-house PA + external buckets all on \$ 1 b).
# 
# ---
# 
# ## 4. Return Equations
# 
# We simulate, for each month \(t\):
# 
# \[
# (r_{\beta,t},\,r_{H,t},\,r_{E,t},\,r_{M,t}) 
# \;\sim\;\text{MVN}\bigl([\mu_{\beta},\,\mu_H,\,\mu_E,\,\mu_M],\,\Sigma\bigr),
# \]
# with
# - \(\mu_{\beta} = \mu_{\text{idx}}\) (monthly mean from analysis window),  
# - \(\mu_H = \frac{\mu_H^{(\text{annual})}}{12}\),  
# - \(\mu_E = \frac{\mu_E^{(\text{annual})}}{12}\),  
# - \(\mu_M = \frac{\mu_M^{(\text{annual})}}{12}\).  
# 
# Covariance \(\Sigma\) built from:  
# - \(\sigma_{\beta} = \sigma_{\text{ref}}\) (monthly vol from reference window),  
# - \(\sigma_H = \sigma_H^{(\text{annual})}/\sqrt{12}\),  
# - \(\sigma_E = \sigma_E^{(\text{annual})}/\sqrt{12}\),  
# - \(\sigma_M = \sigma_M^{(\text{annual})}/\sqrt{12}\),  
# - Pairwise correlations \(\rho_{\beta,H},\,\rho_{\beta,E},\,\rho_{\beta,M},\,\rho_{H,E},\,\dots\).  
# 
# Additionally, each month we draw a financing cost:
# \[
# f_t = \frac{\text{financing_mean}}{12} + \varepsilon_t,\quad
# \varepsilon_t \sim \mathcal{N}\bigl(0,\;(\tfrac{\text{financing_vol}}{12})^2\bigr),
# \]
# with probability \(\text{spike_prob}\) of a jump \(=\text{spike_factor} \times \frac{\text{financing_vol}}{12}\).
# 
# ---
# 
# ### 4.1. Base (All In-House) Strategy
# 
# \[
# R_{\text{Base},t}
# = \; (r_{\beta,t} - f_t)\,\times\,w_{\beta_H}
# \;+\; r_{H,t}\,\times\,w_{\alpha_H}.
# \]
# By default, \(w_{\beta_H} = 0.50\) and \(w_{\alpha_H} = 0.50\).
# 
# ---
# 
# ### 4.2. External PA Strategy
# 
# - Capital allocated: \(X = \text{external_pa_capital}\).  
# - Manager buys \$ X m of index (β) and allocates \(\theta_{\mathrm{ExtPA}} = \text{external_pa_alpha_frac}\) of that \$ X m to α.  
# 
# Return formula:
# \[
# R_{\text{ExtPA},t}
# = \underbrace{\frac{X}{1000}}_{w_{\beta}^{\text{ExtPA}}}\,(r_{\beta,t} - f_t)
# \;+\;\underbrace{\tfrac{X}{1000} \,\times\,\theta_{\mathrm{ExtPA}}}_{w_{\alpha}^{\text{ExtPA}}}\;(r_{M,t}).
# \]
# - If \(\theta_{\mathrm{ExtPA}} = 0.50\), then half of \$ X m is alpha, half is index.
# 
# ---
# 
# ### 4.3. Active Extension Strategy
# 
# - Capital allocated: \(Y = \text{active_ext_capital}\).  
# - Manager runs a long/short portfolio with **Active share** \(S = \frac{\text{active_share_percent}}{100}\).  
#   - E.g. 150/50 → \(S = 0.50\).  
#   - 170/70 → \(S = 0.70\).
# 
# Return formula:
# \[
# R_{\text{ActExt},t}
# = \underbrace{\frac{Y}{1000}}_{w_{\beta}^{\text{ActExt}}}\,(r_{\beta,t} - f_t)
# \;+\;\underbrace{\frac{Y}{1000}\,\times\,S}_{w_{\alpha}^{\text{ActExt}}}\;(r_{E,t}).
# \]
# - The manager’s long/short is embedded in \(r_{E,t}\).  
# 
# ---
# 
# ### 4.4. Internal Margin & Internal PA
# 
# Because both external PA and active-extension managers hold their own index exposure, on your books you only need to hold margin for a single \$ 1 b of index. That is:
# \[
# W = \sigma_{\text{ref}} \times (\mathrm{sd\_of\_vol\_mult}) \times 1000 \quad (\text{\$ mm}).
# \]
# Then you also decide to run \(Z = \text{internal_pa_capital}\) in-house PA:
# 
# - **Internal Beta (margin):**  
#   \[
#   R_{\text{IntBet},t}
#   = \Bigl(\tfrac{W}{1000}\Bigr)\,(r_{\beta,t} - f_t).
#   \]
# - **Internal PA alpha:**  
#   \[
#   R_{\text{IntPA},t}
#   = \Bigl(\tfrac{Z}{1000}\Bigr)\,(r_{H,t}).
#   \]
# - **Internal cash leftover:**  
#   \[
#   \text{internal\_cash\_leftover} = 1000 - W - Z \quad (\text{if positive, earns 0}).
#   \]
# 
# ---
# 
# ## 5. Putting It All Together in Simulation
# 
# 1. **Read user inputs** (via `load_parameters()`):
#    - Dates: `start_date`, `end_date`, `ref_start_date`, `ref_end_date`
#    - Vol/risk: `sd_of_vol_mult`
#    - Returns: `financing_mean`, `financing_vol`, `μ_H`, `σ_H`, `μ_E`, `σ_E`, `μ_M`, `σ_M`
#    - Correlations: `ρ_{β,H}`, `ρ_{β,E}`, `ρ_{β,M}`, `ρ_{H,E}`, `ρ_{H,M}`, `ρ_{E,M}`
#    - Capital buckets: `external_pa_capital`, `external_pa_alpha_frac`, `active_ext_capital`, `active_share_percent`, `internal_pa_capital`
#    - Total fund capital (mm): default = 1000
# 
# 2. **Load index CSV** → `idx_full` (monthly total returns).
# 
# 3. **Filter**  
#    - **`idx_series`** = `idx_full[ start_date : end_date ]` → used for μ_β and σ_β.  
#    - **`idx_ref`** = `idx_full[ ref_start_date : ref_end_date ]` → used for σ_ref.
# 
# 4. **Compute**  
#    \[
#      \mu_{\beta} = \mathrm{mean}(idx\_series), 
#      \quad
#      \sigma_{\beta} = \mathrm{std}(idx\_series),
#      \quad
#      \sigma_{\text{ref}} = \mathrm{std}(idx\_ref).
#    \]
# 
# 5. **Margin-backing**  
#    \[
#      W = \sigma_{\text{ref}} \times \mathrm{sd\_of\_vol\_mult} \times 1000.
#    \]
#    If \(W + Z > 1000\), error. Else compute
#    \[
#      \text{internal\_cash\_leftover} = 1000 - W - Z.
#    \]
# 
# 6. **Build covariance matrix** \(\Sigma\) for \((r_{\beta}, r_H, r_E, r_M)\) using  
#    \(\sigma_{\beta} = \sigma_{\text{ref}},\; \sigma_H = \frac{\sigma_H^{(\text{annual})}}{\sqrt{12}},\; \sigma_E = \frac{\sigma_E^{(\text{annual})}}{\sqrt{12}},\; \sigma_M = \frac{\sigma_M^{(\text{annual})}}{\sqrt{12}},\)  
#    and correlations.
# 
# 7. **Monte Carlo draws**:  
#    For each of \(N_{\text{SIMULATIONS}}\) trials, simulate a \(T=N_{\text{MONTHS}}\)-month path of \(\,(r_{\beta,t},\,r_{H,t},\,r_{E,t},\,r_{M,t})\) and financing \(f_t\).
# 
# 8. **Compute monthly returns** for each bucket:
#    - **Base**:  
#      \[
#        R_{\text{Base},t} 
#        = (r_{\beta,t} - f_t)\,w_{\beta_H} \;+\; r_{H,t}\,w_{\alpha_H}.
#      \]
#    - **External PA**:  
#      \[
#        R_{\text{ExtPA},t} 
#        = \bigl(\tfrac{X}{1000}\bigr)(r_{\beta,t} - f_t) 
#        \;+\; \bigl(\tfrac{X}{1000}\,\theta_{\mathrm{ExtPA}}\bigr)(r_{M,t}).
#      \]
#    - **Active Extension**:  
#      \[
#        R_{\text{ActExt},t} 
#        = \bigl(\tfrac{Y}{1000}\bigr)(r_{\beta,t} - f_t) 
#        \;+\; \bigl(\tfrac{Y}{1000}\,S\bigr)(r_{E,t}).
#      \]
#    - **Internal Beta**:  
#      \[
#        R_{\text{IntBet},t} 
#        = \bigl(\tfrac{W}{1000}\bigr)(r_{\beta,t} - f_t).
#      \]
#    - **Internal PA α**:  
#      \[
#        R_{\text{IntPA},t} 
#        = \bigl(\tfrac{Z}{1000}\bigr)(r_{H,t}).
#      \]
# 
#    Note: We only report three portfolios—“Base,” “ExternalPA,” and “ActiveExt.” Each one compounds its own monthly returns for a 12-month horizon:
#    \[
#      R_{\text{bucket}}^{\text{(year)}} 
#      = \prod_{t=1}^{12} (1 + R_{\text{bucket},t}) - 1.
#    \]
# 
# 9. **Compute performance metrics** for each portfolio’s annual returns:
#    - **Ann Return** = sample mean.  
#    - **Ann Vol** = sample standard deviation.  
#    - **VaR 95%** = 5th percentile.  
#    - **Tracking Error** = std of (bucket_return − index_return).  
#    - **Breach Probability** = % of months (in the first sim path) where \((r_{\text{bucket},t} < -\,\mathrm{buffer\_multiple}\times\sigma_{\beta})\).
# 
# 10. **Export**  
#     - **Inputs sheet:** all parameters (dates, vol caps, bucket sizes, α fractions, active share, σ_ref, W, internal cash leftover, etc.).  
#     - **Summary sheet:** metrics for “Base,” “ExternalPA,” and “ActiveExt.”  
#     - **Raw returns sheets:** monthly paths for each bucket (first simulation) so users can inspect breach months.
# 
# ---
# 
# ## 6. Input Parameters Summary
# 
# Below is a consolidated list of every input variable that must appear in the “friendly” CSV:
# 
# 1. **Date ranges**  
#    - `Start date` → `start_date` (analysis window begin).  
#    - `End date` → `end_date` (analysis window end).  
#    - `Reference start date` → `ref_start_date` (for σ_ref).  
#    - `Reference end date` → `ref_end_date` (for σ_ref).  
# 
# 2. **Financing parameters**  
#    - `Annual financing mean (%)` → `financing_mean_annual` (default = 0.50 %).  
#    - `Annual financing vol (%)` → `financing_vol_annual` (default = 0.10 %).  
#    - `Monthly spike probability` → `spike_prob` (default = 2 %).  
#    - `Spike size (σ × multiplier)` → `spike_factor` (default = 2.25).  
# 
# 3. **In-House PA parameters**  
#    - `In-House annual return (%)` → `mu_H` (default = 4.00 %).  
#    - `In-House annual vol (%)` → `sigma_H` (default = 1.00 %).  
#    - `In-House β` → `w_beta_H` (default = 0.50).  
#    - `In-House α` → `w_alpha_H` (default = 0.50).  
# 
# 4. **Extension α parameters**  
#    - `Alpha-Extension annual return (%)` → `mu_E` (default = 5.00 %).  
#    - `Alpha-Extension annual vol (%)` → `sigma_E` (default = 2.00 %).  
#    - `Active Extension capital (mm)` → `active_ext_capital` (default = 0).  
#    - `Active share (%)` → `active_share_percent` (default = 50 % ⇒ a 150/50 program).  
# 
# 5. **External PA α parameters**  
#    - `External annual return (%)` → `mu_M` (default = 3.00 %).  
#    - `External annual vol (%)` → `sigma_M` (default = 2.00 %).  
#    - `External PA capital (mm)` → `external_pa_capital` (default = 0).  
#    - `External PA α fraction (%)` → `external_pa_alpha_frac` (default = 50 %).  
# 
# 6. **Correlations**  
#    - `Corr index–In-House` → `rho_idx_H` (default = 0.05).  
#    - `Corr index–Alpha-Extension` → `rho_idx_E` (default = 0.00).  
#    - `Corr index–External` → `rho_idx_M` (default = 0.00).  
#    - `Corr In-House–Alpha-Extension` → `rho_H_E` (default = 0.10).  
#    - `Corr In-House–External` → `rho_H_M` (default = 0.10).  
#    - `Corr Alpha-Extension–External` → `rho_E_M` (default = 0.00).  
# 
# 7. **Capital & risk backing**  
#    - `Total fund capital (mm)` → `total_fund_capital` (default = 1000).  
#    - `Standard deviation multiple` → `sd_of_vol_mult` (default = 3).  
#    - `Internal PA capital (mm)` → `internal_pa_capital` (default = 0).  
#    - `Buffer multiple` → `buffer_multiple` (default = 3).  
# 
# 8. **Legacy/Optional**  
#    - `X grid (mm)` → `X_grid_list` (list of X values).  
#    - `External manager α fractions` → `EM_thetas_list`.
# 
# ---
# 
# ## 7. Output Considerations
# 
# 1. **Inputs sheet (Excel):**  
#    List every single parameter, including:  
#    - Date windows (analysis and reference),  
#    - Financing parameters,  
#    - α-stream parameters,  
#    - Correlations,  
#    - Capital buckets (X, Y, Z),  
#    - SD multiple, margin backing \(W\), internal cash leftover,  
#    - Active share, etc.
# 
# 2. **Summary sheet (Excel):**  
#    For each portfolio (“Base,” “ExternalPA,” “ActiveExt”), show:  
#    - Annual Return (%),  
#    - Annual Volatility (%),  
#    - 95 % VaR (%),  
#    - Tracking Error (%),  
#    - Breach Probability (%).
# 
# 3. **Raw returns sheets (Excel):**  
#    Monthly paths for each bucket (first simulation), so users can inspect “breach” months where \(R_{t} < -(\text{buffer_multiple} × σ_{\beta})\).
# 
# 4. **Console output:**  
#    A “human‐friendly” summary, e.g.:  
#    > For “ExternalPA (X = 300, 50 % α)”:  
#    > • Expected annual return: 10.2 %  
#    > • Annual volatility: 12.3 %  
#    > • 95 % VaR: −3.4 %  
#    > • Tracking error: 8.7 %  
#    > • Breach probability: 2.0 %.
# 
# ---
# 
# ## 8. Intuition Behind Key Pieces
# 
# 1. **Why a separate reference period?**  
#    - If you measure index volatility over the same window you analyze (e.g. 2015–2020), you capture “current regime” vol. Often, managers prefer a longer/different window (e.g. 2010–2014) to gauge typical funding volatility. That reference σₙ, times a multiple (e.g. 3×), tells you how much cash to set aside to back \$ 1 b of index exposure.
# 
# 2. **Why Active share as a percentage?**  
#    - A “150/50” program has 150 % long and 50 % short = net 100 %. Its “active share” is reported as 50 %.  
#    - If you want “170/70,” then active share = 70 %.  
#    - The code converts “Active share (%)” to decimal \(S\). For a 150/50 program, the default is 50 % (\(S = 0.50\)).
# 
# 3. **Why each bucket’s formula ensures no double-counting**  
#    - Whenever you give \$ X m to External PA, that manager holds the index exposure on your behalf. You do not hold margin for that portion. Similarly, the Active Extension manager holds their own index.  
#    - On your books, you only need to hold margin for a single \$ 1 b index. That is \(W\).  
#    - Once you hand \$ X m to external PA and \$ Y m to active ext, **both managers** hold \((X + Y)\) of index on your behalf. So your margin \(W\) backs the *entire* \$ 1 b, not just the “leftover” portion.
# 
# ---
# 
# ## 9. Step-by-Step Implementation Checklist
# 
# 1. **Read and parse user parameters** (dates, vols, α fractions, active share, capital buckets, etc.).  
# 2. **Load index CSV** → `idx_full`.  
# 3. **Filter** → `idx_ref` for σ_ref; `idx_series` for μ_β and σ_β.  
# 4. **Compute**:  
#    \[
#      μ_β = \mathrm{mean}(idx\_series), 
#      \quad
#      σ_β = \mathrm{std}(idx\_series), 
#      \quad
#      σ_{\text{ref}} = \mathrm{std}(idx\_ref).
#    \]
# 5. **Margin-backing**:  
#    \[
#      W = σ_{\text{ref}} × (\mathrm{sd\_of\_vol\_mult}) × 1000.
#    \]
#    Check \(W + Z ≤ 1000\). Compute leftover internal cash = \(1000 - W - Z\).
# 
# 6. **Build covariance matrix** using \((σ_{\text{ref}},\,σ_H/√{12},\,σ_E/√{12},\,σ_M/√{12})\) plus correlations.
# 
# 7. **Monte Carlo draws**:  
#    For each of \(N_{\mathrm{SIM}}\) trials, simulate a path of length \(T = N_{\mathrm{MONTHS}}\) for \((r_{\beta,t},\,r_{H,t},\,r_{E,t},\,r_{M,t})\) and financing \(f_t\).
# 
# 8. **Compute monthly returns**:
#    - **Base**:  
#      \[
#        R_{\text{Base},t} = (r_{\beta,t} - f_t)\,w_{\beta_H} + r_{H,t}\,w_{\alpha_H}.
#      \]
#    - **External PA**:  
#      \[
#        R_{\text{ExtPA},t}
#        = \Bigl(\tfrac{X}{1000}\Bigr)(r_{\beta,t} - f_t)
#        \;+\;\Bigl(\tfrac{X}{1000}\,\theta_{\mathrm{ExtPA}}\Bigr)(r_{M,t}).
#      \]
#    - **Active Extension**:  
#      \[
#        R_{\text{ActExt},t}
#        = \Bigl(\tfrac{Y}{1000}\Bigr)(r_{\beta,t} - f_t)
#        \;+\;\Bigl(\tfrac{Y}{1000}\,S\Bigr)(r_{E,t}).
#      \]
#    - **Internal Beta**:  
#      \[
#        R_{\text{IntBet},t} 
#        = \Bigl(\tfrac{W}{1000}\Bigr)(r_{\beta,t} - f_t).
#      \]
#    - **Internal PA α**:  
#      \[
#        R_{\text{IntPA},t} 
#        = \Bigl(\tfrac{Z}{1000}\Bigr)(r_{H,t}).
#      \]
# 
# 9. **Aggregate monthly → annual returns** for “Base,” “ExternalPA,” “ActiveExt.”  
# 10. **Compute metrics**:  
#     - Ann Return, Ann Vol, VaR 95, Tracking Error, Breach Probability.  
# 11. **Export** Inputs, Summary, Raw returns to Excel + print narrative.
# 
# ---

# In[ ]:




