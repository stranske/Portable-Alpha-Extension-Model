from __future__ import annotations

import base64
import io
import json
import os
from types import SimpleNamespace
from typing import Any, Dict, Mapping, cast

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from ..config import ModelConfig
from ..viz import risk_return, theme

__all__ = ["export_to_excel"]

_ONE_PX_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR4nGNgYAAAAAMAASsJTYQAAAAASUVORK5CYII="
)

_LEGACY_SUMMARY_COLUMN_MAP = {
    "AnnReturn": "terminal_AnnReturn",
    "ExcessReturn": "terminal_ExcessReturn",
    "AnnVol": "monthly_AnnVol",
    "VaR": "monthly_VaR",
    "CVaR": "monthly_CVaR",
    "TerminalCVaR": "terminal_CVaR",
    "MaxDD": "monthly_MaxDD",
    "TimeUnderWater": "monthly_TimeUnderWater",
    "BreachProb": "monthly_BreachProb",
    "BreachCount": "monthly_BreachCountPath0",
    "BreachCountPath0": "monthly_BreachCountPath0",
    "ShortfallProb": "terminal_ShortfallProb",
    "TrackingErr": "monthly_TE",
    "TE": "monthly_TE",
}


def normalize_summary_columns(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of summary columns normalized to monthly_/terminal_ prefixes."""
    if summary_df.empty:
        return summary_df
    rename: dict[str, str] = {}
    for old, new in _LEGACY_SUMMARY_COLUMN_MAP.items():
        if old in summary_df.columns and new not in summary_df.columns:
            rename[old] = new
    if not rename:
        return summary_df
    return summary_df.rename(columns=rename)


def export_to_excel(
    inputs_dict: Dict[str, Any],
    summary_df: pd.DataFrame,
    raw_returns_dict: Dict[str, Any],
    filename: str = "Outputs.xlsx",
    *,
    pivot: bool = False,
    diff_config_df: pd.DataFrame | None = None,
    diff_metrics_df: pd.DataFrame | None = None,
    metadata: Mapping[str, Any] | None = None,
    finalize: bool = True,
) -> None:
    """Write inputs, summary, and raw returns into an Excel workbook.

    Parameters
    ----------
    inputs_dict : dict
        Mapping of input parameter names to values.
    summary_df : pandas.DataFrame
        Summary metrics to write to the ``Summary`` sheet.
    raw_returns_dict : dict[str, pandas.DataFrame]
        Per-agent returns matrices.
    filename : str, optional
        Destination Excel file name. Defaults to ``"Outputs.xlsx"``.
    pivot : bool, optional
        If ``True``, collapse all raw returns into a single ``AllReturns`` sheet
        in long format (``Sim``, ``Month``, ``Agent``, ``Return``). Otherwise a
        separate sheet is written per agent. Defaults to ``False``.
    metadata : Mapping[str, Any], optional
        Optional metadata key-value pairs to include in a ``Metadata`` sheet.
    finalize : bool, optional
        If ``True``, apply formatting and embed charts after writing sheets.
        When appending extra sheets later, set to ``False`` and call
        ``finalize_excel_workbook`` after the append. Defaults to ``True``.
    """

    attr_df = _optional_df(inputs_dict, "_attribution_df")
    sleeve_attr_df = _optional_df(inputs_dict, "_sleeve_attribution_df")
    risk_df = _optional_df(inputs_dict, "_risk_attr_df")
    trade_df = _optional_df(inputs_dict, "_tradeoff_df")
    constraint_df = _optional_df(inputs_dict, "_constraint_report_df")
    agent_semantics_df = _ensure_agent_semantics_df(inputs_dict)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_inputs = pd.DataFrame(
            {
                "Parameter": list(inputs_dict.keys()),
                "Value": list(inputs_dict.values()),
            }
        )
        df_inputs.to_excel(writer, sheet_name="Inputs", index=False)
        if metadata:
            meta_df = pd.DataFrame(
                {
                    "Key": list(metadata.keys()),
                    "Value": [_serialize_metadata_value(v) for v in metadata.values()],
                }
            )
            meta_df.to_excel(writer, sheet_name="Metadata", index=False)
        summary_df = normalize_summary_columns(summary_df.copy())
        summary_df["terminal_ShortfallProb"] = summary_df.get(
            "terminal_ShortfallProb", theme.DEFAULT_SHORTFALL_PROB
        )
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Optional: Sensitivity sheet if provided
        sens_df = _optional_df(inputs_dict, "_sensitivity_df")
        if sens_df is not None and not sens_df.empty:
            # Write a concise view
            cols = [
                c
                for c in [
                    "Parameter",
                    "Base",
                    "Minus",
                    "Plus",
                    "Low",
                    "High",
                    "DeltaAbs",
                ]
                if c in sens_df.columns
            ]
            sens_df[cols].to_excel(writer, sheet_name="Sensitivity", index=False)
        # Optional diff sheets
        if diff_config_df is not None and not diff_config_df.empty:
            diff_config_df.to_excel(writer, sheet_name="ConfigDiff", index=False)
        if diff_metrics_df is not None and not diff_metrics_df.empty:
            diff_metrics_df.to_excel(writer, sheet_name="MetricDiff", index=False)

        # Optional: write Attribution sheet if provided in inputs_dict
        if attr_df is not None and not attr_df.empty:
            cols = [c for c in ["Agent", "Sub", "Return"] if c in attr_df.columns]
            if cols:
                attr_df[cols].to_excel(writer, sheet_name="Attribution", index=False)

        # Optional: write sleeve attribution sheet if provided
        if sleeve_attr_df is not None and not sleeve_attr_df.empty:
            cols = [
                c
                for c in ["Agent", "ReturnContribution", "CVaRContribution"]
                if c in sleeve_attr_df.columns
            ]
            if cols:
                sleeve_attr_df[cols].to_excel(writer, sheet_name="sleeve_attribution", index=False)

        # Optional: write RiskAttribution sheet if provided
        if risk_df is not None and not risk_df.empty:
            cols = [
                c
                for c in [
                    "Agent",
                    "BetaVol",
                    "AlphaVol",
                    "CorrWithIndex",
                    "AnnVolApprox",
                    "TEApprox",
                ]
                if c in risk_df.columns
            ]
            if cols:
                risk_df[cols].to_excel(writer, sheet_name="RiskAttribution", index=False)

        # Optional: write Sleeve Trade-offs sheet if provided in inputs_dict
        if trade_df is not None and not trade_df.empty:
            trade_df.to_excel(writer, sheet_name="SleeveTradeoffs", index=True)
        if constraint_df is not None and not constraint_df.empty:
            constraint_df.to_excel(writer, sheet_name="ConstraintBreaches", index=False)
        if agent_semantics_df is not None and not agent_semantics_df.empty:
            agent_semantics_df.to_excel(writer, sheet_name="AgentSemantics", index=False)

        # Write returns either pivoted or per-sheet
        if pivot:
            frames = []
            for name, df in raw_returns_dict.items():
                if isinstance(df, pd.DataFrame) and not df.empty:
                    stacked = df.stack()
                    stacked.name = "Return"
                    long_df = stacked.reset_index()
                    long_df.columns = ["Sim", "Month", "Return"]
                    long_df["Agent"] = name
                    frames.append(long_df[["Sim", "Month", "Agent", "Return"]])
            if frames:
                all_returns = pd.concat(frames, ignore_index=True)
                all_returns.to_excel(writer, sheet_name="AllReturns", index=False)
        else:
            for sheet_name, df in raw_returns_dict.items():
                if isinstance(df, pd.DataFrame):
                    safe_name = sheet_name if len(sheet_name) <= 31 else sheet_name[:31]
                    df.to_excel(writer, sheet_name=safe_name, index=True)

    if finalize:
        finalize_excel_workbook(
            filename,
            inputs_dict,
            summary_df,
        )


def _optional_df(inputs_dict: Dict[str, Any], key: str) -> pd.DataFrame | None:
    value = inputs_dict.get(key)
    return value if isinstance(value, pd.DataFrame) else None


def _coerce_agent_semantics_df(value: Any) -> pd.DataFrame | None:
    if isinstance(value, pd.DataFrame):
        return value
    if isinstance(value, (list, tuple, dict)):
        try:
            return pd.DataFrame(value)
        except Exception:
            return None
    return None


def _ensure_agent_semantics_df(inputs_dict: Dict[str, Any]) -> pd.DataFrame | None:
    value = inputs_dict.get("_agent_semantics_df")
    df = _coerce_agent_semantics_df(value)
    if isinstance(df, pd.DataFrame) and not df.empty:
        return df
    agents = inputs_dict.get("agents")
    total_capital = inputs_dict.get("total_fund_capital")
    if agents is None or total_capital is None:
        return df if isinstance(df, pd.DataFrame) else None
    try:
        from .agent_semantics import build_agent_semantics
    except Exception:
        return df if isinstance(df, pd.DataFrame) else None
    cfg = SimpleNamespace(agents=agents, total_fund_capital=total_capital)
    for attr in (
        "reference_sigma",
        "volatility_multiple",
        "financing_model",
        "financing_schedule_path",
        "financing_term_months",
    ):
        if attr in inputs_dict:
            setattr(cfg, attr, inputs_dict[attr])
    df = build_agent_semantics(cast(ModelConfig, cfg))
    inputs_dict["_agent_semantics_df"] = df.to_dict(orient="records")
    return df


def _serialize_metadata_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, sort_keys=True)
    return value


def finalize_excel_workbook(
    filename: str, inputs_dict: Dict[str, Any], summary_df: pd.DataFrame
) -> None:
    """Apply formatting and embed charts once all sheets are written."""
    summary_df = normalize_summary_columns(summary_df.copy())
    sens_df = _optional_df(inputs_dict, "_sensitivity_df")
    attr_df = _optional_df(inputs_dict, "_attribution_df")

    wb = openpyxl.load_workbook(filename)
    if inputs_dict.get("correlation_repair_applied"):
        existing = wb.properties.keywords or ""
        tag = "correlation_repair_applied=true"
        if tag not in existing:
            wb.properties.keywords = f"{existing}; {tag}".strip("; ")
    max_autosize_cells = 50_000
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        # Avoid expensive autosizing on large sheets (e.g., raw returns).
        if ws.max_row * ws.max_column > max_autosize_cells:
            continue
        for column_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            col_idx = cast(int, column_cells[0].column)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    if "Summary" in wb.sheetnames and not (
        os.environ.get("CI") or os.environ.get("PYTEST_CURRENT_TEST")
    ):
        summary_df = summary_df.copy()
        summary_df["terminal_ShortfallProb"] = summary_df.get(
            "terminal_ShortfallProb", theme.DEFAULT_SHORTFALL_PROB
        )
        ws = wb["Summary"]
        metrics = {
            "terminal_AnnReturn",
            "monthly_AnnVol",
            "monthly_VaR",
            "monthly_CVaR",
            "terminal_CVaR",
            "monthly_BreachProb",
            "monthly_TE",
        }
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for idx, col_name in enumerate(header, 1):
            if col_name in metrics:
                col_letter = get_column_letter(idx)
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.00%"

        try:
            img_bytes = risk_return.make(summary_df).to_image(format="png", engine="kaleido")
            img = XLImage(io.BytesIO(img_bytes))
            ws.add_image(img, "H2")
        except (KeyError, ValueError, RuntimeError, OSError, MemoryError):
            # Some tests pass a minimal summary without expected columns like 'Agent' or 'monthly_AnnVol'; skip chart.
            pass

    # Best-effort: embed tornado image on Sensitivity sheet
    if "Sensitivity" in wb.sheetnames:
        try:
            from ..viz import tornado

            ws = wb["Sensitivity"]
            series: pd.Series | None = None
            if sens_df is not None and not sens_df.empty:
                if {"Parameter", "DeltaAbs"} <= set(sens_df.columns):
                    series = tornado.series_from_sensitivity(sens_df)
            if series is None:
                # Build figure from the written sheet as a fallback
                values: Any = ws.values
                df = pd.DataFrame(values)
                df.columns = df.iloc[0]
                df = df.drop(index=0)
                param_col = "Parameter" if "Parameter" in df.columns else df.columns[0]
                value_col = "DeltaAbs" if "DeltaAbs" in df.columns else df.columns[1]
                series = df.set_index(param_col)[value_col].astype(float)
            fig = tornado.make(cast(pd.Series, series))
            if os.environ.get("CI") or os.environ.get("PYTEST_CURRENT_TEST"):
                img_bytes = _ONE_PX_PNG
            else:
                img_bytes = fig.to_image(format="png", engine="kaleido")
            img = XLImage(io.BytesIO(img_bytes))
            ws.add_image(img, "H2")
        except Exception:
            # Non-fatal if renderer or data missing
            pass

    # Best-effort: embed sunburst image on Attribution sheet
    if "Attribution" in wb.sheetnames:
        try:
            from ..viz import sunburst

            ws = wb["Attribution"]
            # Use attr_df directly instead of reconstructing from worksheet
            if attr_df is not None and not attr_df.empty:
                # Ensure required columns exist
                if {"Agent", "Sub", "Return"} <= set(attr_df.columns):
                    if os.environ.get("CI") or os.environ.get("PYTEST_CURRENT_TEST"):
                        img_bytes = _ONE_PX_PNG
                    else:
                        fig = sunburst.make(attr_df)
                        img_bytes = fig.to_image(format="png", engine="kaleido")
                    img = XLImage(io.BytesIO(img_bytes))
                    ws.add_image(img, "H2")
        except Exception:
            pass

    wb.save(filename)
