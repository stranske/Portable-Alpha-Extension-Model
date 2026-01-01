from __future__ import annotations

import io
import os
from typing import Iterable

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from ..contracts import SUMMARY_REQUIRED_COLUMNS
from ..types import SweepResult
from ..viz import risk_return, theme

__all__ = ["export_sweep_results"]

_SUMMARY_COLUMNS = [*SUMMARY_REQUIRED_COLUMNS, "Combination"]


def export_sweep_results(results: Iterable[SweepResult], filename: str = "Sweep.xlsx") -> None:
    """Write sweep results to an Excel workbook with one sheet per combination."""
    all_summary: pd.DataFrame | None = None

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary_frames = []
        for res in results:
            sheet = f"Run{res['combination_id']}"
            summary_obj = res["summary"]
            if not isinstance(summary_obj, pd.DataFrame):
                continue
            summary = summary_obj.copy()
            summary["ShortfallProb"] = summary.get("ShortfallProb", theme.DEFAULT_SHORTFALL_PROB)
            summary.to_excel(writer, sheet_name=sheet, index=False)
            summary["Combination"] = sheet
            summary_frames.append(summary)

        if summary_frames:
            combined = pd.concat(summary_frames, ignore_index=True)
            combined.to_excel(writer, sheet_name="Summary", index=False)
            all_summary = combined
        else:
            empty_summary = pd.DataFrame(columns=_SUMMARY_COLUMNS)
            empty_summary.to_excel(writer, sheet_name="Summary", index=False)
            all_summary = empty_summary

    wb = openpyxl.load_workbook(filename)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            col_idx = column_cells[0].column
            if col_idx is not None:  # Type guard for mypy/pyright
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    if (
        "Summary" in wb.sheetnames
        and all_summary is not None
        and not all_summary.empty
        and not (os.environ.get("CI") or os.environ.get("PYTEST_CURRENT_TEST"))
    ):
        ws = wb["Summary"]
        metrics = {"AnnReturn", "AnnVol", "VaR", "BreachProb", "TE"}
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for idx, col_name in enumerate(header, 1):
            if col_name in metrics:
                col_letter = get_column_letter(idx)
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.00%"

        try:
            fig = risk_return.make(all_summary)
            img_bytes = fig.to_image(format="png", engine="kaleido")
            img = XLImage(io.BytesIO(img_bytes))
            ws.add_image(img, "H2")
        except (AttributeError, ValueError, RuntimeError, OSError, MemoryError):
            pass

    wb.save(filename)
